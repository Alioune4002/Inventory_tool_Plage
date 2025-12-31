from datetime import timedelta, datetime
import difflib
import unicodedata

import os
import requests

from rest_framework import status, viewsets, permissions, exceptions
from rest_framework.decorators import api_view, permission_classes, renderer_classes
from rest_framework.response import Response
from rest_framework.renderers import BaseRenderer

from django.http import JsonResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import numbers
import re
import logging
import csv
import io
import json
import hashlib
from django.db import transaction
from django.db.models import Q, F, Count, Case, When, IntegerField, Sum
from django.core.cache import cache
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.text import slugify
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import HexColor
from reportlab.graphics.barcode import code128

from accounts.mixins import TenantQuerySetMixin
from accounts.models import Service
from accounts.utils import get_tenant_for_request, get_service_from_request, get_user_role
from accounts.permissions import ProductPermission, ManagerPermission
from accounts.services.access import (
    check_entitlement,
    check_limit,
    get_usage,
    get_retention_days,
    get_limits,
    get_entitlements,
    LimitExceeded,
)
from utils.sendgrid_email import send_email_with_sendgrid
from utils.renderers import XLSXRenderer, CSVRenderer
from inventory.metrics import track_export_event, track_off_lookup_failure

from .models import (
    Product,
    Category,
    LossEvent,
    ExportEvent,
    CatalogPdfEvent,
    LabelPdfEvent,
    Supplier,
    Receipt,
    ReceiptLine,
    ReceiptImportEvent,
    ProductMergeLog,
)
from .serializers import ProductSerializer, CategorySerializer, LossEventSerializer
from .sku import generate_auto_sku

logger = logging.getLogger(__name__)

EXPORT_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
EXPORT_STRIPE_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
EXPORT_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
EXPORT_HEADER_FONT = Font(bold=True, color="FFFFFF")
OFF_TIMEOUT_SECONDS = 3
OFF_LOG_CODE = "OFF_LOOKUP_FAILED"
OFF_CACHE_TTL_SECONDS = 60 * 60 * 48
CATALOG_PDF_MAX_PRODUCTS = 500
LABELS_PDF_MAX_PRODUCTS = 400
RECEIPT_IMPORT_MAX_LINES = 500
RECEIPTS_OCR_PROVIDER_DEFAULT = "ocrspace"
RECEIPTS_OCR_MIN_TEXT_LEN = 30  # en dessous => on considère que le PDF est scanné / vide
RECEIPTS_OCR_MAX_PAGES_DEFAULT = 3
RECEIPTS_OCR_TIMEOUT_DEFAULT = 25
RECEIPTS_OCR_CACHE_TTL_DEFAULT = 60 * 60 * 24  # 24h
DUPLICATE_NAME_SIMILARITY = 0.985
CATALOG_TEMPLATES = {
   
    "classic":  {"accent": "#1E3A8A", "muted": "#475569"},
    "midnight": {"accent": "#0F172A", "muted": "#334155"},
    "emerald":  {"accent": "#047857", "muted": "#475569"},

    # nouveaux templates (premium + variés)
    "royal":    {"accent": "#4F46E5", "muted": "#475569", "header_bg": "#0B1220", "header_text": "#FFFFFF"},
    "sunset":   {"accent": "#F97316", "muted": "#7C2D12", "header_bg": "#111827", "header_text": "#FFFFFF"},
    "rose":     {"accent": "#E11D48", "muted": "#64748B", "header_bg": "#0B1220", "header_text": "#FFFFFF"},
    "ocean":    {"accent": "#0284C7", "muted": "#475569", "header_bg": "#082F49", "header_text": "#FFFFFF"},
    "lime":     {"accent": "#65A30D", "muted": "#475569", "header_bg": "#0B1220", "header_text": "#FFFFFF"},
    "amber":    {"accent": "#D97706", "muted": "#475569", "header_bg": "#111827", "header_text": "#FFFFFF"},
    "violet":   {"accent": "#7C3AED", "muted": "#475569", "header_bg": "#0B1220", "header_text": "#FFFFFF"},
    "copper":   {"accent": "#B45309", "muted": "#57534E", "header_bg": "#0B1220", "header_text": "#FFFFFF"},

    # sobres / industries
    "slate":    {"accent": "#334155", "muted": "#475569", "header_bg": "#0F172A", "header_text": "#FFFFFF"},
    "graphite": {"accent": "#111827", "muted": "#6B7280", "header_bg": "#0B1220", "header_text": "#FFFFFF"},
    "sand":     {"accent": "#A16207", "muted": "#57534E", "header_bg": "#1C1917", "header_text": "#FFFFFF"},
    "ink":      {"accent": "#0EA5E9", "muted": "#64748B", "header_bg": "#020617", "header_text": "#FFFFFF"},

    # beauty / luxe
    "champagne":{"accent": "#C8A97E", "muted": "#6B7280", "header_bg": "#111827", "header_text": "#FFFFFF"},
    "noir_gold":{"accent": "#D4AF37", "muted": "#9CA3AF", "header_bg": "#050505", "header_text": "#FFFFFF"},
    "plum":     {"accent": "#6D28D9", "muted": "#6B7280", "header_bg": "#1F2937", "header_text": "#FFFFFF"},

    # nature / bio
    "forest":   {"accent": "#166534", "muted": "#475569", "header_bg": "#052E16", "header_text": "#FFFFFF"},
    "mint":     {"accent": "#10B981", "muted": "#475569", "header_bg": "#0B1220", "header_text": "#FFFFFF"},
    "earth":    {"accent": "#92400E", "muted": "#57534E", "header_bg": "#1C1917", "header_text": "#FFFFFF"},
}
NAME_STOPWORDS = {
    "produit",
    "article",
    "test",
    "item",
    "lot",
    "pack",
    "piece",
    "pieces",
    "pcs",
    "kg",
    "g",
    "gr",
    "l",
    "ml",
    "cl",
    "x",
    "packaging",
}
LABEL_ALLOWED_FIELDS = {
    "price",
    "price_unit",
    "tva",
    "supplier",
    "brand",
    "unit",
    "dlc",
}
CATALOG_ALLOWED_FIELDS = {
    "barcode",
    "sku",
    "unit",
    "purchase_price",
    "selling_price",
    "tva",
    "variants",
    "dlc",
    "lot",
    "min_qty",
    "supplier",
    "notes",
    "brand",
}
CATALOG_DEFAULT_FIELDS = ["barcode", "sku", "unit"]
CATALOG_FIELD_LABELS = {
    "barcode": "EAN",
    "sku": "SKU",
    "unit": "Unité",
    "purchase_price": "Prix achat",
    "selling_price": "Prix vente",
    "tva": "TVA",
    "variants": "Variante",
    "dlc": "DLC",
    "lot": "Lot",
    "min_qty": "Stock min",
    "supplier": "Fournisseur",
    "notes": "Notes",
    "brand": "Marque",
}


def _is_numeric(value):
    return isinstance(value, numbers.Number) and not isinstance(value, bool)


def _parse_positive_int(value, default, max_value=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    if parsed <= 0:
        return default
    if max_value is not None:
        return min(parsed, max_value)
    return parsed


def _retention_start(tenant):
    days = get_retention_days(tenant)
    if not days:
        return None
    return timezone.now() - timedelta(days=days)


def _apply_retention(qs, tenant):
    start = _retention_start(tenant)
    if not start:
        return qs
    return qs.filter(created_at__gte=start)


def _track_off_failure(reason):
    try:
        day = timezone.now().strftime("%Y-%m-%d")
        key = f"off_lookup_errors:{day}"
        count = int(cache.get(key) or 0) + 1
        cache.set(key, count, OFF_CACHE_TTL_SECONDS)
        track_off_lookup_failure(reason)
        return count
    except Exception:
        logger.warning("OFF_LOOKUP_METRIC_FAILED code=OFF_LOOKUP_METRIC_FAILED reason=%s", reason)
        return None


def _month_start(dt):
    return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def _export_limit_for_format(tenant, export_format):
    limits = get_limits(tenant)
    if export_format == "csv":
        return limits.get("csv_monthly_limit")
    return limits.get("xlsx_monthly_limit")


def _enforce_export_quota(tenant, export_format):
    limit = _export_limit_for_format(tenant, export_format)
    if limit is None:
        return
    start = _month_start(timezone.now())
    used = ExportEvent.objects.filter(
        tenant=tenant,
        format=export_format,
        created_at__gte=start,
    ).count()
    if used >= int(limit):
        code = "LIMIT_EXPORT_CSV_MONTH" if export_format == "csv" else "LIMIT_EXPORT_XLSX_MONTH"
        raise LimitExceeded(code=code, detail="Limite d’export mensuelle atteinte pour votre plan.")


def _log_export_event(tenant, user, export_format, emailed=False, params=None):
    try:
        ExportEvent.objects.create(
            tenant=tenant,
            user=user if user and getattr(user, "is_authenticated", False) else None,
            format=export_format,
            emailed=bool(emailed),
            params=params or {},
        )
    except Exception:
        logger.exception("ExportEvent log failed")
    try:
        track_export_event(export_format, emailed)
    except Exception:
        logger.exception("ExportEvent metric failed")


def _pdf_catalog_limit(tenant):
    limits = get_limits(tenant)
    return limits.get("pdf_catalog_monthly_limit")


def _enforce_pdf_catalog_quota(tenant):
    limit = _pdf_catalog_limit(tenant)
    if limit is None:
        return
    start = _month_start(timezone.now())
    used = CatalogPdfEvent.objects.filter(tenant=tenant, created_at__gte=start).count()
    if used >= int(limit):
        raise LimitExceeded(code="LIMIT_PDF_CATALOG_MONTH", detail="Limite mensuelle du catalogue PDF atteinte.")


def _log_catalog_pdf_event(tenant, user, params=None):
    try:
        CatalogPdfEvent.objects.create(
            tenant=tenant,
            user=user if user and getattr(user, "is_authenticated", False) else None,
            params=params or {},
        )
    except Exception:
        logger.exception("CatalogPdfEvent log failed")


def _labels_pdf_limit(tenant):
    limits = get_limits(tenant)
    return limits.get("labels_pdf_monthly_limit")


def _enforce_labels_pdf_quota(tenant):
    limit = _labels_pdf_limit(tenant)
    if limit is None:
        return
    start = _month_start(timezone.now())
    used = LabelPdfEvent.objects.filter(tenant=tenant, created_at__gte=start).count()
    if used >= int(limit):
        raise LimitExceeded(code="LIMIT_LABELS_PDF_MONTH", detail="Limite mensuelle d’étiquettes PDF atteinte.")


def _log_labels_pdf_event(tenant, user, params=None):
    try:
        LabelPdfEvent.objects.create(
            tenant=tenant,
            user=user if user and getattr(user, "is_authenticated", False) else None,
            params=params or {},
        )
    except Exception:
        logger.exception("LabelPdfEvent log failed")


def _receipts_import_limit(tenant):
    limits = get_limits(tenant)
    return limits.get("receipts_import_monthly_limit")


def _enforce_receipts_import_quota(tenant):
    limit = _receipts_import_limit(tenant)
    if limit is None:
        return
    start = _month_start(timezone.now())
    used = ReceiptImportEvent.objects.filter(tenant=tenant, created_at__gte=start).count()
    if used >= int(limit):
        raise LimitExceeded(code="LIMIT_RECEIPTS_IMPORT_MONTH", detail="Limite mensuelle d’imports atteinte.")


def _log_receipt_import_event(tenant, user, params=None):
    try:
        ReceiptImportEvent.objects.create(
            tenant=tenant,
            user=user if user and getattr(user, "is_authenticated", False) else None,
            params=params or {},
        )
    except Exception:
        logger.exception("ReceiptImportEvent log failed")


def _parse_pdf_fields(raw_fields):
    if not raw_fields:
        return list(CATALOG_DEFAULT_FIELDS)
    fields = [f.strip().lower() for f in str(raw_fields).split(",") if f.strip()]
    filtered = [f for f in fields if f in CATALOG_ALLOWED_FIELDS]
    return filtered if filtered else list(CATALOG_DEFAULT_FIELDS)


def _format_catalog_field(label, value):
    if value is None or value == "":
        return None
    return f"{label}: {value}"


def _format_catalog_line(product, fields, include_service=False, currency_code="EUR"):
    values = []
    if include_service and getattr(product, "service", None):
        values.append(_format_catalog_field("Service", product.service.name))

    if "barcode" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["barcode"], product.barcode))
    if "sku" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["sku"], product.internal_sku))
    if "unit" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["unit"], product.unit))
    if "variants" in fields:
        variant = _format_variant(product)
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["variants"], variant))
    if "purchase_price" in fields:
        val = _format_price(product.purchase_price, currency_code)
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["purchase_price"], val))
    if "selling_price" in fields:
        val = _format_price(product.selling_price, currency_code)
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["selling_price"], val))
    if "tva" in fields:
        val = f"{product.tva}%" if product.tva is not None else None
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["tva"], val))
    if "dlc" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["dlc"], product.dlc))
    if "lot" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["lot"], product.lot_number))
    if "min_qty" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["min_qty"], product.min_qty))
    if "supplier" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["supplier"], product.supplier))
    if "brand" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["brand"], product.brand))
    if "notes" in fields:
        values.append(_format_catalog_field(CATALOG_FIELD_LABELS["notes"], product.notes))

    return " · ".join([v for v in values if v])


def _currency_symbol(code: str) -> str:
    code = (code or "EUR").upper()
    symbols = {
        "EUR": "€",
        "USD": "$",
        "GBP": "£",
        "CHF": "CHF",
        "CAD": "CA$",
        "AUD": "A$",
        "JPY": "¥",
        "CNY": "¥",
        "BRL": "R$",
        "MAD": "MAD",
        "XOF": "F CFA",
        "XAF": "F CFA",
    }
    return symbols.get(code, code)


def _format_price(value, currency_code="EUR"):
    if value is None:
        return None
    try:
        symbol = _currency_symbol(currency_code)
        return f"{float(value):.2f} {symbol}"
    except (TypeError, ValueError):
        return None


def _price_per_unit(product, currency_code="EUR"):
    price = product.selling_price if product.selling_price is not None else product.purchase_price
    if price is None:
        return None, None
    unit = (product.unit or "").lower()
    try:
        price = float(price)
    except (TypeError, ValueError):
        return None, None
    symbol = _currency_symbol(currency_code)
    if unit == "kg":
        return price, f"{symbol}/kg"
    if unit == "g":
        return price * 1000, f"{symbol}/kg"
    if unit == "l":
        return price, f"{symbol}/L"
    if unit == "ml":
        return price * 1000, f"{symbol}/L"
    return None, None


def _parse_promo(raw_type, raw_value):
    promo_type = (raw_type or "").strip().lower()
    if promo_type not in ("percent", "amount"):
        return None
    try:
        value = float(raw_value)
    except (TypeError, ValueError):
        return None
    if promo_type == "percent" and not (0 < value <= 100):
        return None
    if promo_type == "amount" and value <= 0:
        return None
    return {"type": promo_type, "value": value}


def _apply_promo(price, promo):
    if price is None or not promo:
        return None
    try:
        base = float(price)
    except (TypeError, ValueError):
        return None
    if promo["type"] == "percent":
        new_price = base * (1 - promo["value"] / 100.0)
    else:
        new_price = base - promo["value"]
    return max(new_price, 0)


def _format_label_date(product):
    if not product.dlc:
        return None, None
    label = "DLC"
    if product.expiry_type and product.expiry_type != "none":
        label = str(product.expiry_type)
    return label, product.dlc.strftime("%d/%m/%Y")


def _format_pack_info(product):
    if product.pack_size and product.pack_uom:
        try:
            return f"{float(product.pack_size):.3f} {product.pack_uom}"
        except (TypeError, ValueError):
            return f"{product.pack_size} {product.pack_uom}"
    if product.conversion_factor and product.conversion_unit:
        try:
            return f"{float(product.conversion_factor):.3f} {product.conversion_unit}"
        except (TypeError, ValueError):
            return f"{product.conversion_factor} {product.conversion_unit}"
    return None


def _parse_label_counts(raw):
    if not raw:
        return {}
    counts = {}
    for part in str(raw).split(","):
        if ":" not in part:
            continue
        pid, count = part.split(":", 1)
        if not pid.strip().isdigit():
            continue
        try:
            count_val = int(count)
        except (TypeError, ValueError):
            continue
        if count_val < 1:
            continue
        counts[int(pid)] = min(count_val, 50)
    return counts


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def _draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.setFillColorRGB(0.45, 0.45, 0.45)
        self.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Page {self._pageNumber} / {page_count}")


class PDFRenderer(BaseRenderer):
    media_type = "application/pdf"
    format = "pdf"
    charset = None

    def render(self, data, accepted_media_type=None, renderer_context=None):
        if data is None:
            return b""
        response = renderer_context.get("response") if renderer_context else None
        if response is not None and response.status_code >= 400:
            response["Content-Type"] = "application/json"
            return json.dumps(data).encode("utf-8")
        if isinstance(data, bytes):
            return data
        if isinstance(data, str):
            return data.encode()
        try:
            return bytes(data)
        except TypeError:
            return json.dumps(data).encode("utf-8")

def _parse_ids_param(raw):
    if not raw:
        return []
    if isinstance(raw, (list, tuple)):
        vals = raw
    else:
        vals = str(raw).split(",")
    out = []
    for v in vals:
        v = str(v).strip()
        if v.isdigit():
            out.append(int(v))
    return out


def _draw_circle_logo(c, img_bytes, x, y, size):
    """
    Draw a circular clipped logo (premium look).
    x,y = bottom-left of the square container.
    """
    if not img_bytes:
        return
    try:
        logo = ImageReader(io.BytesIO(img_bytes))
        path = c.beginPath()
        r = size / 2.0
        cx = x + r
        cy = y + r
        path.circle(cx, cy, r)
        c.saveState()
        c.clipPath(path, stroke=0, fill=0)
        c.drawImage(
            logo,
            x,
            y,
            width=size,
            height=size,
            preserveAspectRatio=True,
            mask="auto",
        )
        c.restoreState()

        # subtle ring
        c.saveState()
        c.setLineWidth(1)
        c.setStrokeColor(HexColor("#E2E8F0"))
        c.circle(cx, cy, r)
        c.restoreState()
    except Exception:

        return


def _truncate_text(c, text, font_name, font_size, max_width):
    if not text:
        return ""
    s = str(text)
    if c.stringWidth(s, font_name, font_size) <= max_width:
        return s
    ell = "…"
    lo, hi = 0, len(s)
    while lo < hi:
        mid = (lo + hi) // 2
        cand = s[:mid].rstrip() + ell
        if c.stringWidth(cand, font_name, font_size) <= max_width:
            lo = mid + 1
        else:
            hi = mid
    return (s[: max(0, lo - 1)].rstrip() + ell) if lo > 0 else ell
def _build_catalog_pdf(
    *,
    tenant,
    company_name,
    company_email,
    company_phone,
    company_address,
    fields,
    products,
    truncated,
    include_service,
    logo_bytes=None,
    template="classic",
):
    buffer = io.BytesIO()
    c = NumberedCanvas(buffer, pagesize=A4)
    width, height = A4

    theme = CATALOG_TEMPLATES.get((template or "").lower(), CATALOG_TEMPLATES["classic"])

    accent = HexColor(theme.get("accent", "#1E3A8A"))
    muted = HexColor(theme.get("muted", "#475569"))

    header_bg = HexColor(theme.get("header_bg", "#0B1220"))      
    header_text = HexColor(theme.get("header_text", "#FFFFFF")) 

    ink = HexColor("#0F172A")
    paper = HexColor("#FFFFFF")
    soft = HexColor("#F1F5F9")
    border = HexColor("#E2E8F0")

    margin_x = 1.6 * cm
    top_header_h = 2.2 * cm

    def draw_top_header(title_left=None):
        # Dark premium bar
        c.setFillColor(header_bg)
        c.rect(0, height - top_header_h, width, top_header_h, fill=1, stroke=0)

        # Accent line
        c.setFillColor(accent)
        c.rect(0, height - top_header_h, width, 0.10 * cm, fill=1, stroke=0)

        # Circular logo
        if logo_bytes:
            _draw_circle_logo(c, logo_bytes, margin_x, height - top_header_h + 0.35 * cm, 1.5 * cm)

        # Company name
        c.setFillColor(header_text)
        c.setFont("Helvetica-Bold", 14)
        name_x = margin_x + (1.75 * cm if logo_bytes else 0)
        c.drawString(name_x, height - 1.35 * cm, _truncate_text(c, company_name or tenant.name or "StockScan", "Helvetica-Bold", 14, width - name_x - margin_x))

        # subtitle
        c.setFont("Helvetica", 9.5)
        subtitle = title_left or "Catalogue produits"
        c.setFillColor(HexColor(theme.get("header_subtext", "#CBD5E1")))
        c.drawString(name_x, height - 1.90 * cm, _truncate_text(c, subtitle, "Helvetica", 9.5, width - name_x - margin_x))

        # date right
        c.setFont("Helvetica", 9.5)
        c.setFillColor(HexColor(theme.get("header_subtext", "#CBD5E1")))
        c.drawRightString(width - margin_x, height - 1.85 * cm, timezone.now().strftime("%d/%m/%Y"))

    def draw_cover():
        # Background
        c.setFillColor(paper)
        c.rect(0, 0, width, height, fill=1, stroke=0)

        draw_top_header("Catalogue produits")

        # Hero title
        c.setFillColor(ink)
        c.setFont("Helvetica-Bold", 22)
        c.drawString(margin_x, height - 4.2 * cm, "Catalogue produits")

        c.setFont("Helvetica", 11)
        c.setFillColor(muted)
        c.drawString(margin_x, height - 4.95 * cm, "Export premium StockScan")

        # Info card
        card_x = margin_x
        card_y = height - 8.9 * cm
        card_w = width - 2 * margin_x
        card_h = 3.2 * cm

        c.setFillColor(soft)
        c.setStrokeColor(border)
        c.setLineWidth(1)
        c.roundRect(card_x, card_y, card_w, card_h, 10, fill=1, stroke=1)

        info_lines = [company_email, company_phone, company_address]
        y = card_y + card_h - 0.75 * cm
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(ink)
        c.drawString(card_x + 0.7 * cm, y, "Coordonnées")
        y -= 0.55 * cm

        c.setFont("Helvetica", 9.5)
        c.setFillColor(muted)
        for line in info_lines:
            if line:
                c.drawString(card_x + 0.7 * cm, y, str(line))
                y -= 0.5 * cm

        if truncated:
            c.setFillColor(HexColor("#B91C1C"))
            c.setFont("Helvetica-Bold", 9.5)
            c.drawString(margin_x, card_y - 0.65 * cm, f"Liste tronquée à {CATALOG_PDF_MAX_PRODUCTS} produits.")

    # Cover
    draw_cover()
    c.showPage()

    # Group by category
    grouped = {}
    for p in products:
        key = (p.category or "Sans catégorie").strip() or "Sans catégorie"
        grouped.setdefault(key, []).append(p)

    # Layout for product cards: 2 columns
    col_gap = 0.7 * cm
    card_w = (width - 2 * margin_x - col_gap) / 2
    card_h = 4.0 * cm
    y = height - top_header_h - 1.0 * cm

    def new_page():
        c.showPage()
        draw_top_header("Catalogue produits")
        return height - top_header_h - 1.0 * cm

    draw_top_header("Catalogue produits")

    for category, items in grouped.items():
        if y < 3.2 * cm:
            y = new_page()

        # Category pill
        c.setFillColor(HexColor("#0B1220"))
        pill_h = 0.85 * cm
        c.roundRect(margin_x, y - pill_h, width - 2 * margin_x, pill_h, 10, fill=1, stroke=0)
        c.setFillColor(paper)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin_x + 0.55 * cm, y - 0.62 * cm, _truncate_text(c, category, "Helvetica-Bold", 11, width - 2 * margin_x - 1.2 * cm))
        y -= (pill_h + 0.55 * cm)

        col = 0
        x = margin_x

        for p in items:
            if y < 3.2 * cm:
                y = new_page()
                # repeat category label lightly
                c.setFillColor(muted)
                c.setFont("Helvetica-Bold", 10)
                c.drawString(margin_x, y, _truncate_text(c, category, "Helvetica-Bold", 10, width - 2 * margin_x))
                y -= 0.55 * cm

            # Card background
            c.setFillColor(soft)
            c.setStrokeColor(border)
            c.setLineWidth(1)
            c.roundRect(x, y - card_h, card_w, card_h, 12, fill=1, stroke=1)

            # Name
            c.setFillColor(ink)
            c.setFont("Helvetica-Bold", 11)
            name = (p.name or "").strip()
            c.drawString(x + 0.55 * cm, y - 0.70 * cm, _truncate_text(c, name, "Helvetica-Bold", 11, card_w - 1.1 * cm))

            # Service tag
            if include_service and getattr(p, "service", None):
                tag = (p.service.name or "").strip()
                c.setFont("Helvetica", 8.5)
                c.setFillColor(accent)
                c.drawString(x + 0.55 * cm, y - 1.25 * cm, _truncate_text(c, f"Service: {tag}", "Helvetica", 8.5, card_w - 1.1 * cm))

            # Details line(s)
            details = _format_catalog_line(p, fields, include_service=False, currency_code=tenant.currency_code)
            c.setFont("Helvetica", 9)
            c.setFillColor(muted)
            if details:
                # allow 2 lines
                max_w = card_w - 1.1 * cm
                line1 = _truncate_text(c, details, "Helvetica", 9, max_w)
                c.drawString(x + 0.55 * cm, y - 1.85 * cm, line1)
                if len(details) > len(line1):
                    rest = details[len(line1) :].lstrip(" ·")
                    if rest:
                        line2 = _truncate_text(c, rest, "Helvetica", 9, max_w)
                        c.drawString(x + 0.55 * cm, y - 2.35 * cm, line2)

            # Barcode small (optional)
            code = (p.barcode or p.internal_sku or "").strip()
            if code:
                try:
                    barcode_h = 0.75 * cm
                    barcode_obj = code128.Code128(code, barHeight=barcode_h, barWidth=0.65)
                    bx = x + card_w - 0.55 * cm - barcode_obj.width
                    by = y - card_h + 0.55 * cm
                    barcode_obj.drawOn(c, bx, by)
                    c.setFont("Helvetica", 7)
                    c.setFillColor(muted)
                    c.drawRightString(x + card_w - 0.55 * cm, by - 0.25 * cm, _truncate_text(c, code, "Helvetica", 7, card_w * 0.6))
                except Exception:
                    pass

            # Move column / row
            col += 1
            if col % 2 == 0:
                # new row
                x = margin_x
                y -= (card_h + 0.6 * cm)
            else:
                # next column
                x = margin_x + card_w + col_gap

        # ensure next category starts on new row properly
        if (len(items) % 2) == 1:
            x = margin_x
            y -= (card_h + 0.6 * cm)
        else:
            x = margin_x

        y -= 0.2 * cm

    c.save()
    return buffer.getvalue()


def _converted_quantity(product):
    factor = getattr(product, "conversion_factor", None)
    unit = getattr(product, "conversion_unit", None)
    if factor is None or not unit:
        return None, None
    try:
        return float(product.quantity or 0) * float(factor), unit
    except (TypeError, ValueError):
        return None, unit


def _format_variant(product):
    name = (getattr(product, "variant_name", None) or "").strip()
    value = (getattr(product, "variant_value", None) or "").strip()
    if name and value:
        return f"{name}: {value}"
    return value or name or ""


def _tokenize_name(value):
    if not value:
        return []
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    tokens = [t for t in text.split() if t and t not in NAME_STOPWORDS]
    return tokens


def _normalize_name(value):
    tokens = _tokenize_name(value)
    return " ".join(tokens)


def _similarity(a, b):
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(a=a, b=b).ratio()


def _token_similarity(tokens_a, tokens_b):
    if not tokens_a or not tokens_b:
        return 0.0
    set_a, set_b = set(tokens_a), set(tokens_b)
    if not set_a or not set_b:
        return 0.0
    return len(set_a & set_b) / len(set_a | set_b)


def _name_is_generic(tokens):
    if len(tokens) < 2:
        return True
    text = " ".join(tokens)
    if len(text) < 8:
        return True
    if len(tokens) < 3 and all(len(t) <= 3 for t in tokens):
        return True
    return False


def _tokens_are_specific(tokens):
    if len(tokens) < 2:
        return False
    long_tokens = [t for t in tokens if len(t) >= 4]
    return len(long_tokens) >= 2


def _supporting_signal_count(a, b):
    def same_field(field):
        av = (getattr(a, field, None) or "").strip().lower()
        bv = (getattr(b, field, None) or "").strip().lower()
        return bool(av and bv and av == bv)

    count = 0
    count += 1 if same_field("category") else 0
    count += 1 if same_field("brand") else 0
    count += 1 if same_field("supplier") else 0
    count += 1 if same_field("pack_uom") else 0
    count += 1 if same_field("pack_size") else 0

    try:
        ap = float(a.purchase_price) if a.purchase_price is not None else None
        bp = float(b.purchase_price) if b.purchase_price is not None else None
    except (TypeError, ValueError):
        ap, bp = None, None
    if ap is not None and bp is not None and ap > 0 and bp > 0 and abs(ap - bp) < 0.01:
        count += 1

    return count


def _has_meaningful_signal(a, b):
    def same_field(field):
        av = (getattr(a, field, None) or "").strip().lower()
        bv = (getattr(b, field, None) or "").strip().lower()
        return bool(av and bv and av == bv)

    if same_field("category") or same_field("brand") or same_field("supplier"):
        return True
    if same_field("pack_uom") or same_field("pack_size"):
        return True
    try:
        ap = float(a.purchase_price) if a.purchase_price is not None else None
        bp = float(b.purchase_price) if b.purchase_price is not None else None
    except (TypeError, ValueError):
        ap, bp = None, None
    return ap is not None and bp is not None and ap > 0 and bp > 0 and abs(ap - bp) < 0.01


def _has_core_signal(a, b):
    def same_field(field):
        av = (getattr(a, field, None) or "").strip().lower()
        bv = (getattr(b, field, None) or "").strip().lower()
        return bool(av and bv and av == bv)

    return bool(same_field("category") or same_field("brand") or same_field("supplier"))


def _group_has_core_signal(group):
    if len(group) < 2:
        return False
    for idx, p in enumerate(group):
        for q in group[idx + 1 :]:
            if _has_core_signal(p, q):
                return True
    return False


def _has_strong_signal(a, b):
    return bool(
        (a.barcode and b.barcode and a.barcode == b.barcode)
        or (a.internal_sku and b.internal_sku and a.internal_sku == b.internal_sku)
    )


def _has_supporting_signal(a, b, min_signals=3):
    if _has_strong_signal(a, b):
        return True
    if not _has_meaningful_signal(a, b):
        return False
    return _supporting_signal_count(a, b) >= min_signals


def _group_has_supporting_signal(group, min_signals=3):
    if len(group) < 2:
        return False
    for idx, p in enumerate(group):
        for q in group[idx + 1 :]:
            if _has_supporting_signal(p, q, min_signals=min_signals):
                return True
    return False


def _history_month_cutoff(months):
    if months is None:
        return None
    total = _parse_positive_int(months, None)
    if not total:
        return None
    today = timezone.now().date().replace(day=1)
    year = today.year
    month = today.month - (total - 1)
    while month <= 0:
        month += 12
        year -= 1
    return f"{year:04d}-{month:02d}"


def _build_csv_bytes(headers, rows, delimiter=";", title=None):
    buffer = io.StringIO()
    buffer.write("sep=;\n")
    writer = csv.writer(buffer, delimiter=delimiter)
    if title:
        writer.writerow([title] + [""] * (len(headers) - 1))
    writer.writerow(headers)
    for row in rows:
        writer.writerow(["" if value is None else value for value in row])
    content = buffer.getvalue()
    return ("\ufeff" + content).encode("utf-8")


def _apply_sheet_style(ws, headers, title=None):
    header_row = 1
    if title:
        ws.insert_rows(1)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 26
        header_row = 2

    ws.freeze_panes = f"A{header_row + 1}"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    ws.row_dimensions[header_row].height = 22

    header_lookup = {index + 1: (header or "") for index, header in enumerate(headers)}

    for col_idx, header in header_lookup.items():
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = EXPORT_HEADER_FONT
        cell.fill = EXPORT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = EXPORT_BORDER

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=ws.max_column):
        row_idx = row[0].row
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = EXPORT_STRIPE_FILL
        for cell in row:
            cell.border = EXPORT_BORDER
            header = header_lookup.get(cell.column, "")
            if _is_numeric(cell.value):
                header_lower = str(header).lower()
                if any(term in header_lower for term in ("prix", "valeur", "sale", "purchase", "€")):
                    cell.number_format = "#,##0.00"
                elif any(term in header_lower for term in ("tva", "vat")):
                    cell.number_format = "0.00"
                else:
                    cell.number_format = "#,##0.###"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_cells in ws.columns:
        max_length = 0
        column_letter = None
        for cell in col_cells:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                column_letter = cell.column_letter
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def _format_service_label(tenant, service_param=None):
    if isinstance(service_param, Service):
        return service_param.name
    if service_param and service_param != "all":
        svc = Service.objects.filter(id=service_param, tenant=tenant).first()
        return svc.name if svc else "Service"
    names = list(tenant.services.values_list("name", flat=True))
    if not names:
        return "Tous services"
    if len(names) <= 3:
        return ", ".join(names)
    return f"{', '.join(names[:3])} +{len(names) - 3}"


def _format_period_label(month=None, from_month=None, to_month=None):
    def format_month(val):
        if not val:
            return ""
        match = re.match(r"^(\\d{4})-(\\d{2})$", str(val))
        if match:
            year, mm = match.groups()
            return f"{mm}/{year}"
        return str(val)

    if month:
        return format_month(month)
    if from_month and to_month:
        return f"{format_month(from_month)} → {format_month(to_month)}"
    if from_month:
        return f"Depuis {format_month(from_month)}"
    if to_month:
        return f"Jusqu’à {format_month(to_month)}"
    return "Toutes périodes"


def _build_export_title(tenant, period_label, service_label):
    base = f"Inventaire {period_label}" if period_label else "Inventaire"
    return f"{base} - {tenant.name} (Services : {service_label})"


def _build_export_filename(prefix, period_label, tenant_name, service_label, ext):
    parts = [prefix, period_label, tenant_name, service_label]
    safe = [slugify(p) for p in parts if p]
    if not safe:
        safe = [prefix]
    return f"{'_'.join(safe)}.{ext}"


def home(request):
    return JsonResponse({"message": "Bienvenue sur l'API de Inventory Tool ! Utilisez /api/products/ pour accéder aux données."})


class ProductViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated, ProductPermission]

    def get_queryset(self):
        qs = super().get_queryset()
        tenant = get_tenant_for_request(self.request)
        service = get_service_from_request(self.request)
        qs = qs.filter(service=service)

        month = self.request.query_params.get("month")
        if month:
            qs = qs.filter(inventory_month=month)

        qs = _apply_retention(qs, tenant)
        return qs

    def perform_create(self, serializer):
        role = get_user_role(self.request)
        if role not in ["owner", "manager", "operator"]:
            raise exceptions.PermissionDenied("Rôle insuffisant pour créer un produit.")
        tenant = get_tenant_for_request(self.request)
        service = get_service_from_request(self.request)
        usage = get_usage(tenant)
        check_limit(tenant, "max_products", usage["products_count"], requested_increment=1)
        serializer.save(tenant=tenant, service=service)

    def perform_update(self, serializer):
        role = get_user_role(self.request)
        if role not in ["owner", "manager", "operator"]:
            raise exceptions.PermissionDenied("Rôle insuffisant pour modifier un produit.")
        tenant = get_tenant_for_request(self.request)
        service = get_service_from_request(self.request)
        serializer.save(tenant=tenant, service=service)

    def perform_destroy(self, instance):
        role = get_user_role(self.request)
        if role not in ["owner", "manager"]:
            raise exceptions.PermissionDenied("Rôle insuffisant pour supprimer un produit.")
        instance.is_archived = True
        instance.archived_at = timezone.now()
        instance.save(update_fields=["is_archived", "archived_at"])


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ProductPermission])
def lookup_product(request):
    barcode = request.query_params.get("barcode", "")
    if not barcode:
        return Response({"detail": "Paramètre barcode requis."}, status=400)

    tenant = get_tenant_for_request(request)
    service = get_service_from_request(request)

    product_qs = Product.objects.filter(tenant=tenant, service=service, barcode=barcode)
    product_qs = _apply_retention(product_qs, tenant)
    product = product_qs.order_by("-inventory_month", "-created_at").first()

    if product:
        serializer = ProductSerializer(product)

        recent = Product.objects.filter(tenant=tenant, service=service).order_by("-created_at")
        recent = _apply_retention(recent, tenant)[:5]
        history = Product.objects.filter(tenant=tenant, service=service, barcode=barcode).order_by(
            "-inventory_month", "-created_at"
        )
        history = _apply_retention(history, tenant)
        history_limit = _parse_positive_int(request.query_params.get("history_limit"), 200, 500)
        history_months = request.query_params.get("history_months")
        if history_months in (None, ""):
            history_months = 12
        history_cutoff = _history_month_cutoff(history_months)
        if history_cutoff:
            history = history.filter(inventory_month__gte=history_cutoff)
        history = history[:history_limit]

        return Response(
            {
                "found": True,
                "product": serializer.data,
                "recent": ProductSerializer(recent, many=True).data,
                "history": ProductSerializer(history, many=True).data,
            }
        )

    if tenant.domain == "food":
        try:
            import socket
            import urllib.error
            import urllib.request

            url = f"https://world.openfoodfacts.org/api/v2/product/{barcode}.json"
            with urllib.request.urlopen(url, timeout=OFF_TIMEOUT_SECONDS) as resp:
                data = json.loads(resp.read().decode())
                if data.get("status") == 1:
                    p = data.get("product", {}) or {}
                    suggestion = {
                        "name": p.get("product_name") or "",
                        "brand": p.get("brands") or "",
                        "category": (p.get("categories_tags") or [None])[0],
                        "quantity": p.get("quantity"),
                    }
                    return Response({"found": False, "suggestion": suggestion})
        except (urllib.error.URLError, urllib.error.HTTPError, socket.timeout) as exc:
            count = _track_off_failure("url_error")
            logger.warning(
                "OFF_LOOKUP_FAILED code=%s reason=url_error count=%s barcode=%s",
                OFF_LOG_CODE,
                count,
                barcode,
                exc_info=exc,
            )
            return Response(
                {
                    "found": False,
                    "off_error": "Préremplissage indisponible (OpenFoodFacts). Réessayez plus tard.",
                },
                status=200,
            )
        except Exception:
            count = _track_off_failure("exception")
            logger.warning(
                "OFF_LOOKUP_FAILED code=%s reason=exception count=%s barcode=%s",
                OFF_LOG_CODE,
                count,
                barcode,
                exc_info=True,
            )
            return Response(
                {
                    "found": False,
                    "off_error": "Préremplissage indisponible (OpenFoodFacts). Réessayez plus tard.",
                },
                status=200,
            )

    return Response({"found": False}, status=200)


class CategoryViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated, ManagerPermission]

    def get_queryset(self):
        qs = super().get_queryset()
        service = get_service_from_request(self.request)
        return qs.filter(service=service)

    def perform_create(self, serializer):
        role = get_user_role(self.request)
        if role not in ["owner", "manager"]:
            raise exceptions.PermissionDenied("Rôle insuffisant pour créer une catégorie.")
        tenant = get_tenant_for_request(self.request)
        service = get_service_from_request(self.request)
        if Category.objects.filter(
            tenant=tenant, service=service, name=serializer.validated_data.get("name")
        ).exists():
            raise exceptions.ValidationError("Cette catégorie existe déjà pour ce service.")
        serializer.save()


class LossEventViewSet(TenantQuerySetMixin, viewsets.ModelViewSet):
    queryset = LossEvent.objects.all()
    serializer_class = LossEventSerializer
    permission_classes = [permissions.IsAuthenticated, ProductPermission]

    def get_queryset(self):
        qs = super().get_queryset()
        tenant = get_tenant_for_request(self.request)
        service = get_service_from_request(self.request)
        qs = qs.filter(service=service)
        month = self.request.query_params.get("month")
        if month:
            qs = qs.filter(inventory_month=month)
        qs = _apply_retention(qs, tenant)
        return qs

    def perform_create(self, serializer):
        role = get_user_role(self.request)
        if role not in ["owner", "manager", "operator"]:
            raise exceptions.PermissionDenied("Rôle insuffisant pour déclarer une perte.")
        tenant = get_tenant_for_request(self.request)
        service = get_service_from_request(self.request)
        check_entitlement(tenant, "loss_management")
        product = serializer.validated_data.get("product")
        if not product:
            raise exceptions.ValidationError({"product": "Ce champ est obligatoire."})

        serializer.save(
            tenant=tenant,
            service=service,
            created_by=self.request.user if self.request.user.is_authenticated else None,
        )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
def inventory_stats(request):
    month = request.query_params.get("month", None)
    tenant = get_tenant_for_request(request)
    service = get_service_from_request(request)
    features = getattr(service, "features", {}) or {}
    item_type_cfg = features.get("item_type", {}) or {}
    item_type_enabled = bool(item_type_cfg.get("enabled", False))

    products = Product.objects.filter(tenant=tenant, service=service)
    if month:
        products = products.filter(inventory_month=month)
    products = _apply_retention(products, tenant)

    losses_qs = LossEvent.objects.filter(tenant=tenant, service=service)
    if month:
        losses_qs = losses_qs.filter(inventory_month=month)
    losses_qs = _apply_retention(losses_qs, tenant)

    loss_by_product = {}
    for l in losses_qs:
        if l.product_id:
            loss_by_product[l.product_id] = loss_by_product.get(l.product_id, 0) + float(l.quantity or 0)

    losses_total_qty = sum(float(l.quantity or 0) for l in losses_qs)
    losses_total_cost = 0
    losses_by_reason = []
    reasons = losses_qs.values_list("reason", flat=True).distinct()
    for r in reasons:
        subset = losses_qs.filter(reason=r)
        qty = sum(float(l.quantity or 0) for l in subset)
        cost = 0
        for l in subset:
            if l.product and l.product.purchase_price:
                cost += float(l.quantity or 0) * float(l.product.purchase_price)
        losses_by_reason.append({"reason": r, "total_qty": qty, "total_cost": cost})
        losses_total_cost += cost

    by_product = []
    for p in products:
        stock_final = float(p.quantity or 0)
        loss_qty = loss_by_product.get(p.id, 0)
        purchase_price = float(p.purchase_price or 0)
        selling_price = float(p.selling_price or 0)
        converted_qty, converted_unit = _converted_quantity(p)

        is_raw_material = item_type_enabled and p.product_role == "raw_material"
        selling_value_current = 0
        if selling_price and not is_raw_material:
            selling_value_current = selling_price * stock_final

        by_product.append(
            {
                "name": p.name,
                "category": p.category,
                "unit": p.unit,
                "stock_final": stock_final,
                "losses_qty": loss_qty,
                "product_role": p.product_role,
                "purchase_price": purchase_price,
                "selling_price": selling_price,
                "purchase_value_current": purchase_price * stock_final if purchase_price else 0,
                "selling_value_current": selling_value_current,
                "converted_quantity": converted_qty,
                "converted_unit": converted_unit,
                "quantity_sold_est": None,
                "ca_estime": None,
                "cout_matiere": None,
                "marge": None,
                "notes": (
                    ["Calculs complets limités: stock_initial/entrées/sorties non suivis."]
                    + (["Matière première: valeur de vente potentielle non comptée."] if is_raw_material else [])
                ),
            }
        )

    categories = products.values_list("category", flat=True).distinct()
    by_category = []
    for cat in categories:
        cat_products = products.filter(category=cat)
        total_qty = sum(float(p.quantity or 0) for p in cat_products)
        total_purchase_value = sum((float(p.purchase_price or 0) * float(p.quantity or 0)) for p in cat_products)
        total_selling_value = 0
        for p in cat_products:
            if item_type_enabled and p.product_role == "raw_material":
                continue
            total_selling_value += float(p.selling_price or 0) * float(p.quantity or 0)
        cat_losses = sum(loss_by_product.get(p.id, 0) for p in cat_products)
        by_category.append(
            {
                "category": cat,
                "total_quantity": total_qty,
                "total_purchase_value": total_purchase_value,
                "total_selling_value": total_selling_value,
                "losses_qty": cat_losses,
            }
        )

    total_purchase_value = sum((float(p.purchase_price or 0) * float(p.quantity or 0)) for p in products)
    total_selling_value = 0
    for p in products:
        if item_type_enabled and p.product_role == "raw_material":
            continue
        total_selling_value += float(p.selling_price or 0) * float(p.quantity or 0)

    converted_totals = {}
    for p in products:
        converted_qty, converted_unit = _converted_quantity(p)
        if converted_qty is None or not converted_unit:
            continue
        converted_totals.setdefault(converted_unit, 0)
        converted_totals[converted_unit] += converted_qty

    return Response(
        {
            "total_value": total_purchase_value,
            "total_selling_value": total_selling_value,
            "service_totals": {
                "purchase_value": total_purchase_value,
                "selling_value": total_selling_value,
                "losses_qty": losses_total_qty,
                "losses_cost": losses_total_cost,
                "notes": (
                    [
                        "Quantité vendue et marge estimée limitées : pas de stock initial/entrées/sorties enregistrés.",
                        "Si le module Matière première / Produit fini est actif, la valeur de vente exclut les matières premières.",
                    ]
                    if item_type_enabled
                    else ["Quantité vendue et marge estimée limitées : pas de stock initial/entrées/sorties enregistrés."]
                ),
            },
            "by_product": by_product,
            "by_category": by_category,
            "losses_total_qty": losses_total_qty,
            "losses_total_cost": losses_total_cost,
            "losses_by_reason": losses_by_reason,
            "timeseries": [],
            "categories": by_category,
            "converted_totals": converted_totals,
        }
    )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
@renderer_classes([XLSXRenderer, CSVRenderer])
def export_excel(request):
    month = request.query_params.get("month", None)
    if not month:
        return Response({"error": "Paramètre month requis"}, status=400)

    tenant = get_tenant_for_request(request)
    service = get_service_from_request(request)
    check_entitlement(tenant, "exports_basic")
    _enforce_export_quota(tenant, "xlsx")

    products = Product.objects.filter(tenant=tenant, service=service, inventory_month=month)
    products = _apply_retention(products, tenant)

    headers = [
        "Nom",
        "Catégorie",
        "Prix achat (€)",
        "Prix vente (€)",
        "TVA (%)",
        "DLC",
        "Quantité",
        "Variante",
        "Stock min",
        "Quantité convertie",
        "Unité convertie",
    ]
    rows = []
    for p in products:
        converted_qty, converted_unit = _converted_quantity(p)
        rows.append(
            [
                p.name,
                p.category or "",
                float(p.purchase_price or 0),
                float(p.selling_price or 0),
                float(p.tva or 0) if p.tva is not None else "",
                p.dlc or "",
                float(p.quantity or 0),
                _format_variant(p),
                float(p.min_qty) if p.min_qty is not None else "",
                converted_qty if converted_qty is not None else "",
                converted_unit or "",
            ]
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Inventaire {month}"
    service_label = _format_service_label(tenant, service)
    period_label = _format_period_label(month=month)
    title = _build_export_title(tenant, period_label, service_label)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _apply_sheet_style(ws, headers, title=title)

    bio = io.BytesIO()
    wb.save(bio)
    xlsx_bytes = bio.getvalue()

    filename = _build_export_filename("inventaire", period_label, tenant.name, service_label, "xlsx")
    resp = Response(xlsx_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    _log_export_event(
        tenant=tenant,
        user=request.user,
        export_format="xlsx",
        emailed=False,
        params={"month": month, "service": service.id},
    )
    return resp


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
@renderer_classes([XLSXRenderer, CSVRenderer])
def export_generic(request):
    tenant = get_tenant_for_request(request)

    service_param = request.query_params.get("service")
    export_format = (request.query_params.get("format", "xlsx") or "xlsx").lower()
    mode = request.query_params.get("mode", "all")
    from_month = request.query_params.get("from")
    to_month = request.query_params.get("to")
    email_to = request.query_params.get("email")
    email_message = request.query_params.get("message", "")

    check_entitlement(tenant, "exports_basic")
    if export_format == "xlsx":
        check_entitlement(tenant, "exports_xlsx")
    if email_to:
        check_entitlement(tenant, "exports_email")
    _enforce_export_quota(tenant, export_format)

    qs = Product.objects.filter(tenant=tenant)
    qs = _apply_retention(qs, tenant)
    if service_param and service_param != "all":
        try:
            qs = qs.filter(service_id=int(service_param))
        except ValueError:
            pass

    if from_month:
        qs = qs.filter(inventory_month__gte=from_month)
    if to_month:
        qs = qs.filter(inventory_month__lte=to_month)

    if mode != "all":
        qs = qs.filter(container_status=mode.upper())

    headers = [
        "Month",
        "Service",
        "Category",
        "ProductName",
        "Variant",
        "ContainerStatus",
        "Qty",
        "UOM",
        "MinQty",
        "ConvertedQty",
        "ConvertedUnit",
        "RemainingFraction",
        "PackSize",
        "PackUOM",
        "PurchasePrice",
        "SalePrice",
    ]
    rows = []
    for p in qs.select_related("service"):
        converted_qty, converted_unit = _converted_quantity(p)
        rows.append(
            [
                p.inventory_month,
                p.service.name if p.service else "",
                p.category or "",
                p.name,
                _format_variant(p),
                p.container_status,
                float(p.quantity or 0),
                p.unit or "",
                float(p.min_qty) if p.min_qty is not None else "",
                converted_qty if converted_qty is not None else "",
                converted_unit or "",
                p.remaining_fraction or "",
                p.pack_size or "",
                p.pack_uom or "",
                float(p.purchase_price or 0),
                float(p.selling_price or 0),
            ]
        )

    service_label = _format_service_label(tenant, service_param)
    period_label = _format_period_label(from_month=from_month, to_month=to_month)
    title = _build_export_title(tenant, period_label, service_label)
    filename = _build_export_filename(
        "inventaire",
        period_label,
        tenant.name,
        service_label,
        "xlsx" if export_format == "xlsx" else "csv",
    )
    mimetype = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if export_format == "xlsx"
        else "text/csv"
    )

    if export_format == "csv":
        attachment_bytes = _build_csv_bytes(headers, rows, title=title)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        _apply_sheet_style(ws, headers, title=title)
        bio = io.BytesIO()
        wb.save(bio)
        attachment_bytes = bio.getvalue()

    if email_to:
        send_email_with_sendgrid(
            to_email=email_to,
            subject="Export StockScan",
            text_body=email_message or "Ci-joint votre export.",
            html_body=email_message or None,
            filename=filename,
            file_bytes=attachment_bytes or b"",
            mimetype=mimetype,
            fallback_to_django=True,
        )

    resp = Response(attachment_bytes, content_type=mimetype if export_format == "xlsx" else f"{mimetype}; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    _log_export_event(
        tenant=tenant,
        user=request.user,
        export_format=export_format,
        emailed=bool(email_to),
        params={
            "service": service_param,
            "from": from_month,
            "to": to_month,
            "mode": mode,
            "email": email_to,
        },
    )
    return resp


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
@renderer_classes([XLSXRenderer, CSVRenderer])
def export_advanced(request):
    tenant = get_tenant_for_request(request)
    service_from_request = get_service_from_request(request)
    data = request.data or {}

    from_month = data.get("from_month") or data.get("fromMonth")
    to_month = data.get("to_month") or data.get("toMonth")

    service_single = data.get("service")
    services = data.get("services")
    if not services:
        if service_single:
            services = [service_single]
        else:
            services = [service_from_request.id]

    categories = data.get("categories") or []
    mode = (data.get("mode") or "all").upper()

    price_min = data.get("price_min", data.get("priceMin"))
    price_max = data.get("price_max", data.get("priceMax"))
    stock_min = data.get("stock_min", data.get("stockMin"))

    include_tva = data.get("include_tva", data.get("includeTVA", True))
    include_dlc = data.get("include_dlc", data.get("includeDLC", True))
    include_sku = data.get("include_sku", data.get("includeSKU", True))
    include_summary = data.get("include_summary", data.get("includeSummary", True))
    include_charts = data.get("include_charts", data.get("includeCharts", False))
    export_format = (data.get("format", data.get("exportFormat", "xlsx")) or "xlsx").lower()

    email_to = (data.get("email") or "").strip()
    email_message = (data.get("message") or "").strip()

    fields = data.get("fields") or data.get("columns")
    if isinstance(fields, str):
        fields = [v.strip() for v in fields.split(",") if v.strip()]

    check_entitlement(tenant, "exports_basic")
    if export_format == "xlsx":
        check_entitlement(tenant, "exports_xlsx")
    if include_summary or include_charts:
        check_entitlement(tenant, "reports_advanced")
    if email_to:
        check_entitlement(tenant, "exports_email")
    _enforce_export_quota(tenant, export_format)

    qs = Product.objects.filter(tenant=tenant)
    qs = _apply_retention(qs, tenant)

    def _none_if_blank(v):
        if v is None:
            return None
        if isinstance(v, str) and v.strip() == "":
            return None
        return v

    price_min = _none_if_blank(price_min)
    price_max = _none_if_blank(price_max)
    stock_min = _none_if_blank(stock_min)

    if services:
        qs = qs.filter(service_id__in=services)
    else:
        qs = qs.filter(service=service_from_request)

    if from_month and to_month:
        qs = qs.filter(inventory_month__gte=from_month, inventory_month__lte=to_month)
    elif from_month:
        qs = qs.filter(inventory_month__gte=from_month)
    elif to_month:
        qs = qs.filter(inventory_month__lte=to_month)

    if categories:
        qs = qs.filter(category__in=categories)

    if mode in ("SEALED", "OPENED"):
        qs = qs.filter(container_status=mode)

    if price_min is not None:
        qs = qs.filter(selling_price__gte=price_min)
    if price_max is not None:
        qs = qs.filter(selling_price__lte=price_max)
    if stock_min is not None:
        qs = qs.filter(quantity__gte=stock_min)

    field_map = {
        "name": ("Nom", lambda p: p.name or ""),
        "category": ("Catégorie", lambda p: p.category or ""),
        "purchase_price": ("Prix achat (€)", lambda p: float(p.purchase_price) if p.purchase_price else 0),
        "selling_price": ("Prix vente (€)", lambda p: float(p.selling_price) if p.selling_price else 0),
        "tva": ("TVA (%)", lambda p: float(p.tva) if p.tva else ""),
        "dlc": ("DLC", lambda p: p.dlc or ""),
        "quantity": ("Quantité", lambda p: p.quantity or 0),
        "identifier": (
            "Code-barres / SKU",
            lambda p: p.internal_sku if getattr(p, "no_barcode", False) else (p.barcode or ""),
        ),
        "inventory_month": ("Mois", lambda p: p.inventory_month or ""),
        "service": ("Service", lambda p: p.service.name if p.service else ""),
        "unit": ("Unité", lambda p: p.unit or ""),
        "min_qty": ("Stock min", lambda p: float(p.min_qty) if p.min_qty is not None else ""),
        "variant_name": ("Variante (libellé)", lambda p: p.variant_name or ""),
        "variant_value": ("Variante (valeur)", lambda p: p.variant_value or ""),
        "variant": ("Variante", lambda p: _format_variant(p)),
        "lot_number": ("Lot", lambda p: p.lot_number or ""),
        "container_status": ("Statut", lambda p: p.container_status or ""),
        "remaining_fraction": ("Reste (fraction)", lambda p: p.remaining_fraction if p.remaining_fraction is not None else ""),
        "conversion_unit": ("Unité conversion", lambda p: p.conversion_unit or ""),
        "conversion_factor": ("Facteur conversion", lambda p: float(p.conversion_factor) if p.conversion_factor else ""),
        "converted_quantity": (
            "Quantité convertie",
            lambda p: (_converted_quantity(p)[0] if _converted_quantity(p)[0] is not None else ""),
        ),
        "converted_unit": ("Unité convertie", lambda p: _converted_quantity(p)[1] or ""),
        "brand": ("Marque", lambda p: p.brand or ""),
        "supplier": ("Fournisseur", lambda p: p.supplier or ""),
        "notes": ("Notes", lambda p: p.notes or ""),
        "product_role": ("Type produit", lambda p: p.product_role or ""),
    }

    selected_fields = []
    if isinstance(fields, list):
        selected_fields = [f for f in fields if f in field_map]

    if selected_fields:
        headers = [field_map[f][0] for f in selected_fields]
    else:
        headers = ["Nom", "Catégorie", "Prix achat (€)", "Prix vente (€)"]
        if include_tva:
            headers.append("TVA (%)")
        if include_dlc:
            headers.append("DLC")
        headers.append("Quantité")
        if include_sku:
            headers.append("Code-barres / SKU")
        headers.extend(["Variante", "Stock min", "Quantité convertie", "Unité convertie"])
        headers.append("Mois")
        headers.append("Service")

    def build_row(product):
        if selected_fields:
            return [field_map[f][1](product) for f in selected_fields]
        row = [
            product.name,
            product.category,
            float(product.purchase_price) if product.purchase_price else 0,
            float(product.selling_price) if product.selling_price else 0,
        ]
        if include_tva:
            row.append(float(product.tva) if product.tva else "")
        if include_dlc:
            row.append(product.dlc or "")
        row.append(product.quantity or 0)
        if include_sku:
            row.append(product.internal_sku if getattr(product, "no_barcode", False) else product.barcode or "")
        converted_qty, converted_unit = _converted_quantity(product)
        row.extend(
            [
                _format_variant(product),
                float(product.min_qty) if product.min_qty is not None else "",
                converted_qty if converted_qty is not None else "",
                converted_unit or "",
            ]
        )
        row.append(product.inventory_month)
        row.append(product.service.name if product.service else "")
        return row

    products_qs = qs.select_related("service")
    rows = [build_row(p) for p in products_qs]

    service_ids = [int(s) for s in services if str(s).isdigit()]
    service_label = _format_service_label(tenant, "all" if not service_ids else None)
    if service_ids:
        service_label = _format_service_label(
            tenant,
            Service.objects.filter(id__in=service_ids, tenant=tenant).first(),
        )
        names = list(Service.objects.filter(id__in=service_ids, tenant=tenant).values_list("name", flat=True))
        if names:
            service_label = ", ".join(names) if len(names) <= 3 else f"{', '.join(names[:3])} +{len(names) - 3}"
    period_label = _format_period_label(from_month=from_month, to_month=to_month)
    title = _build_export_title(tenant, period_label, service_label)
    filename = _build_export_filename(
        "export_avance",
        period_label,
        tenant.name,
        service_label,
        "csv" if export_format == "csv" else "xlsx",
    )
    mimetype = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if export_format != "csv"
        else "text/csv"
    )

    if export_format == "csv":
        attachment_bytes = _build_csv_bytes(headers, rows, title=title)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export avancé"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        _apply_sheet_style(ws, headers, title=title)

        if include_summary or include_charts:
            summary = wb.create_sheet("Synthèse")
            summary["A1"] = "Synthèse export"
            summary["A1"].font = Font(bold=True, size=14)

            products_list = list(products_qs)
            total_products = len(products_list)
            total_qty = sum(float(p.quantity or 0) for p in products_list)
            total_purchase = sum(float(p.purchase_price or 0) * float(p.quantity or 0) for p in products_list)

            total_selling = 0
            for p in products_list:
                features = getattr(p.service, "features", {}) or {}
                item_cfg = features.get("item_type", {}) or {}
                if item_cfg.get("enabled") and p.product_role == "raw_material":
                    continue
                total_selling += float(p.selling_price or 0) * float(p.quantity or 0)

            dlc_count = sum(1 for p in products_list if p.dlc)

            has_item_type_enabled = any(
                (getattr(p.service, "features", {}) or {}).get("item_type", {}).get("enabled")
                for p in products_list
            )

            info_rows = [
                ("Produits", total_products),
                ("Quantité totale", total_qty),
                ("Valeur stock achat (€)", total_purchase),
                ("Valeur stock vente (€)", total_selling),
            ]
            if include_dlc and (not selected_fields or "dlc" in selected_fields):
                info_rows.append(("Produits avec DLC/DDM", dlc_count))
            if has_item_type_enabled:
                info_rows.append(("Note", "Valeur de vente exclut les matières premières."))

            row_cursor = 3
            for label, value in info_rows:
                summary.cell(row=row_cursor, column=1, value=label).font = Font(bold=True)
                summary.cell(row=row_cursor, column=2, value=value)
                row_cursor += 1

            row_cursor += 1
            summary.cell(row=row_cursor, column=1, value="Par catégorie").font = Font(bold=True)
            row_cursor += 1

            headers_summary = ["Catégorie", "Quantité", "Valeur achat (€)", "Valeur vente (€)"]
            for col_num, header in enumerate(headers_summary, 1):
                cell = summary.cell(row=row_cursor, column=col_num, value=header)
                cell.font = EXPORT_HEADER_FONT
                cell.fill = EXPORT_HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = EXPORT_BORDER
            row_cursor += 1

            category_totals = {}
            for p in products_list:
                cat = p.category or "Sans catégorie"
                category_totals.setdefault(cat, {"qty": 0, "purchase": 0, "selling": 0})
                category_totals[cat]["qty"] += float(p.quantity or 0)
                category_totals[cat]["purchase"] += float(p.purchase_price or 0) * float(p.quantity or 0)

                features = getattr(p.service, "features", {}) or {}
                item_cfg = features.get("item_type", {}) or {}
                if not (item_cfg.get("enabled") and p.product_role == "raw_material"):
                    category_totals[cat]["selling"] += float(p.selling_price or 0) * float(p.quantity or 0)

            start_category_row = row_cursor
            for cat, totals in sorted(category_totals.items()):
                summary.cell(row=row_cursor, column=1, value=cat)
                summary.cell(row=row_cursor, column=2, value=totals["qty"])
                summary.cell(row=row_cursor, column=3, value=totals["purchase"])
                summary.cell(row=row_cursor, column=4, value=totals["selling"])
                if row_cursor % 2 == 0:
                    for col_idx in range(1, 5):
                        summary.cell(row=row_cursor, column=col_idx).fill = EXPORT_STRIPE_FILL
                for col_idx in range(1, 5):
                    summary.cell(row=row_cursor, column=col_idx).border = EXPORT_BORDER
                row_cursor += 1

            if include_charts and category_totals:
                chart = BarChart()
                chart.title = "Valeur stock achat par catégorie"
                chart.y_axis.title = "€"
                data_ref = Reference(summary, min_col=3, min_row=start_category_row - 1, max_row=row_cursor - 1)
                categories_ref = Reference(summary, min_col=1, min_row=start_category_row, max_row=row_cursor - 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(categories_ref)
                summary.add_chart(chart, "F4")

        for sheet in wb.worksheets:
            for col_cells in sheet.columns:
                max_length = 0
                column_letter = None
                for cell in col_cells:
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        column_letter = cell.column_letter
                        if cell.value is None:
                            continue
                        max_length = max(max_length, len(str(cell.value)))
                if column_letter:
                    sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)

        bio = io.BytesIO()
        wb.save(bio)
        attachment_bytes = bio.getvalue()

    if email_to:
        try:
            send_email_with_sendgrid(
                to_email=email_to,
                subject="Export StockScan",
                text_body=email_message or "Ci-joint votre export.",
                html_body=email_message or None,
                filename=filename,
                file_bytes=attachment_bytes or b"",
                mimetype=mimetype,
                fallback_to_django=True,
            )
        except Exception:
            logger.exception("Erreur envoi email export_advanced")

    resp = Response(
        attachment_bytes,
        content_type=mimetype if export_format != "csv" else f"{mimetype}; charset=utf-8",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    _log_export_event(
        tenant=tenant,
        user=request.user,
        export_format=export_format,
        emailed=bool(email_to),
        params={
            "services": services,
            "from": from_month,
            "to": to_month,
            "mode": mode,
            "format": export_format,
            "email": email_to,
        },
    )
    return resp


@api_view(["GET", "POST"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
@renderer_classes([PDFRenderer])
def catalog_pdf(request):
    tenant = get_tenant_for_request(request)
    check_entitlement(tenant, "pdf_catalog")
    _enforce_pdf_catalog_quota(tenant)

    payload = request.data if request.method == "POST" else request.query_params
    service_param = payload.get("service")
    ids_list = _parse_ids_param(payload.get("ids"))
    include_service = service_param == "all"

    qs = Product.objects.filter(tenant=tenant)
    qs = _apply_retention(qs, tenant)
    if service_param and service_param != "all":
        try:
            qs = qs.filter(service_id=int(service_param))
        except (ValueError, TypeError):
            pass
    if ids_list:
        qs = qs.filter(id__in=ids_list)
    query = payload.get("q")
    if query:
        qs = qs.filter(
            Q(name__icontains=query)
            | Q(barcode__icontains=query)
            | Q(internal_sku__icontains=query)
            | Q(brand__icontains=query)
            | Q(supplier__icontains=query)
        )

    category = payload.get("category")
    if category:
        qs = qs.filter(category__iexact=category)

    qs = qs.select_related("service").order_by("category", "name")
    raw_products = list(qs[: CATALOG_PDF_MAX_PRODUCTS + 1])
    truncated = len(raw_products) > CATALOG_PDF_MAX_PRODUCTS
    products = raw_products[:CATALOG_PDF_MAX_PRODUCTS]

    fields = _parse_pdf_fields(payload.get("fields"))
    company_name = (payload.get("company_name") or "").strip() or tenant.name
    company_email = (payload.get("company_email") or "").strip()
    company_phone = (payload.get("company_phone") or "").strip()
    company_address = (payload.get("company_address") or "").strip()
    template = (payload.get("template") or "classic").strip().lower()
    if template not in CATALOG_TEMPLATES:
        template = "classic"

    logo_bytes = None
    if request.method == "POST":
        logo_file = request.FILES.get("logo")
        if logo_file:
            logo_bytes = logo_file.read()

    pdf_bytes = _build_catalog_pdf(
        tenant=tenant,
        company_name=company_name,
        company_email=company_email,
        company_phone=company_phone,
        company_address=company_address,
        fields=fields,
        products=products,
        truncated=truncated,
        include_service=include_service,
        logo_bytes=logo_bytes,
        template=template,
    )

    _log_catalog_pdf_event(
        tenant=tenant,
        user=request.user,
        params={
            "service": service_param,
            "fields": fields,
            "q": query,
            "category": category,
            "template": template,
            "logo": bool(logo_bytes),
        },
    )

    resp = Response(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="stockscan_catalogue.pdf"'
    return resp

@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
def catalog_templates(request):
    return Response(
        {
            "templates": [
                {"key": k, "accent": v.get("accent"), "header_bg": v.get("header_bg")}
                for k, v in sorted(CATALOG_TEMPLATES.items())
            ]
        }
    )
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ProductPermission])
def search_products(request):
    tenant = get_tenant_for_request(request)
    query = request.query_params.get("q", "")
    service_param = request.query_params.get("service")
    if service_param == "all":
        qs = Product.objects.filter(tenant=tenant)
    else:
        service = get_service_from_request(request)
        qs = Product.objects.filter(tenant=tenant, service=service)
    qs = _apply_retention(qs, tenant)
    if query:
        qs = qs.filter(
            Q(name__icontains=query)
            | Q(barcode__icontains=query)
            | Q(internal_sku__icontains=query)
        )
    results = list(
        qs.order_by("name")[:10].values(
            "id",
            "name",
            "barcode",
            "internal_sku",
            "category",
            "inventory_month",
            "service_id",
            "service__name",
        )
    )
    return Response(results)


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ProductPermission])
def alerts(request):
    tenant = get_tenant_for_request(request)
    service = get_service_from_request(request)

    entitlements = set(get_entitlements(tenant))
    allow_stock = "alerts_stock" in entitlements
    allow_expiry = "alerts_expiry" in entitlements
    if not allow_stock and not allow_expiry:
        raise LimitExceeded(
            code="FEATURE_NOT_INCLUDED",
            detail="Alertes non incluses dans votre plan. Stock: plan Duo+. Dates: plan Multi+.",
        )

    limit = _parse_positive_int(request.query_params.get("limit"), 50, 200)
    offset = _parse_positive_int(request.query_params.get("offset"), 0)

    alerts_list = []
    now = timezone.now().date()
    expiry_critical_days = 30
    expiry_warning_days = 90

    base_qs = Product.objects.filter(tenant=tenant, service=service)
    base_qs = _apply_retention(base_qs, tenant)

    if allow_stock:
        stock_qs = base_qs.filter(min_qty__isnull=False, quantity__lte=F("min_qty"))
        for p in stock_qs:
            qty = float(p.quantity or 0)
            min_qty = float(p.min_qty or 0)
            severity = "critical" if qty <= 0 else "warning"
            alerts_list.append(
                {
                    "type": "stock_low",
                    "severity": severity,
                    "product_id": p.id,
                    "product_name": p.name,
                    "service_id": service.id,
                    "service_name": service.name,
                    "quantity": qty,
                    "min_qty": min_qty,
                    "message": f"Stock bas (min {min_qty}).",
                }
            )

    dlc_enabled = bool((getattr(service, "features", {}) or {}).get("dlc", {}).get("enabled"))
    if allow_expiry and dlc_enabled:
        expiry_qs = base_qs.filter(dlc__isnull=False).exclude(expiry_type="none")
        for p in expiry_qs:
            days_left = (p.dlc - now).days if p.dlc else None
            if days_left is None or days_left > expiry_warning_days:
                continue
            severity = "critical" if days_left <= expiry_critical_days else "warning"
            alerts_list.append(
                {
                    "type": "expiry",
                    "severity": severity,
                    "product_id": p.id,
                    "product_name": p.name,
                    "service_id": service.id,
                    "service_name": service.name,
                    "dlc": p.dlc.isoformat() if p.dlc else None,
                    "days_left": days_left,
                    "message": f"DLC/DDM proche ({days_left}j).",
                }
            )

    alerts_list.sort(
        key=lambda a: (
            0 if a["severity"] == "critical" else 1,
            a.get("days_left", 9999),
            a.get("product_name", ""),
        )
    )
    total = len(alerts_list)
    paginated = alerts_list[offset : offset + limit]

    return Response({"count": total, "limit": limit, "offset": offset, "results": paginated})


def _product_brief(product):
    return {
        "id": product.id,
        "name": product.name,
        "barcode": product.barcode,
        "internal_sku": product.internal_sku,
        "category": product.category,
        "quantity": float(product.quantity or 0),
        "inventory_month": product.inventory_month,
        "created_at": product.created_at.isoformat() if product.created_at else None,
    }


def _pick_master(candidates):
    if not candidates:
        return None
    scored = []
    for p in candidates:
        score = 0
        score += 3 if p.barcode else 0
        score += 3 if p.internal_sku else 0
        score += 1 if p.category else 0
        score += 1 if p.purchase_price is not None else 0
        score += 1 if p.selling_price is not None else 0
        score += float(p.quantity or 0)
        scored.append((score, p))
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1]


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
def product_duplicates(request):
    tenant = get_tenant_for_request(request)
    check_entitlement(tenant, "anti_duplicates")
    service = get_service_from_request(request)
    month = request.query_params.get("month")

    qs = Product.objects.filter(tenant=tenant, service=service)
    if month:
        qs = qs.filter(inventory_month=month)

    duplicates = []

    for field, label in (("barcode", "barcode"), ("internal_sku", "sku")):
        agg = (
            qs.exclude(**{f"{field}__isnull": True})
            .exclude(**{field: ""})
            .values(field)
            .annotate(cnt=Count("id"))
            .filter(cnt__gt=1)
        )
        for row in agg:
            value = row.get(field)
            group = list(qs.filter(**{field: value}))
            master = _pick_master(group)
            duplicates.append(
                {
                    "type": label,
                    "key": value,
                    "master_id": master.id if master else None,
                    "confidence": 0.98 if label == "barcode" else 0.96,
                    "reason": "Même code-barres" if label == "barcode" else "Même SKU",
                    "products": [_product_brief(p) for p in group],
                }
            )

    name_groups = {}
    name_tokens = {}
    for p in qs:
        tokens = _tokenize_name(p.name)
        if _name_is_generic(tokens) or not _tokens_are_specific(tokens):
            continue
        norm = " ".join(tokens)
        if not norm:
            continue
        name_groups.setdefault(norm, []).append(p)
        name_tokens[norm] = tokens

    for norm, group in name_groups.items():
        if len(group) < 2:
            continue
        if not any((not p.barcode and not p.internal_sku) for p in group):
            continue
        if not _group_has_core_signal(group):
            continue
        if not _group_has_supporting_signal(group, min_signals=4):
            continue
        master = _pick_master(group)
        duplicates.append(
            {
                "type": "name",
                "key": norm,
                "master_id": master.id if master else None,
                "confidence": 0.82,
                "reason": "Nom identique + signaux communs (catégorie, prix, fournisseur…)",
                "products": [_product_brief(p) for p in group],
            }
        )

    # fuzzy pass (lightweight) on groups of similar normalized names
    if len(name_groups) <= 120:
        norms = list(name_groups.keys())
        seen_pairs = set()
        for i, base in enumerate(norms):
            for other in norms[i + 1 :]:
                if (base, other) in seen_pairs or (other, base) in seen_pairs:
                    continue
                tokens_a = name_tokens.get(base, [])
                tokens_b = name_tokens.get(other, [])
                shared = set(tokens_a) & set(tokens_b)
                if len(shared) < 3:
                    continue
                if _token_similarity(tokens_a, tokens_b) < 0.9:
                    continue
                if _similarity(base, other) < DUPLICATE_NAME_SIMILARITY:
                    continue
                merged = list({*name_groups.get(base, []), *name_groups.get(other, [])})
                if not any((not p.barcode and not p.internal_sku) for p in merged):
                    continue
                if not _group_has_core_signal(merged):
                    continue
                has_signal = False
                for idx, p in enumerate(merged):
                    for q in merged[idx + 1 :]:
                        if _has_supporting_signal(p, q, min_signals=4):
                            has_signal = True
                            break
                    if has_signal:
                        break

                if not has_signal:
                    continue
                if len(merged) < 2:
                    continue
                master = _pick_master(merged)
                duplicates.append(
                    {
                        "type": "name_fuzzy",
                        "key": f"{base} ~ {other}",
                        "master_id": master.id if master else None,
                        "confidence": 0.76,
                        "reason": "Nom très proche + signaux forts (catégorie/prix/fournisseur)",
                        "products": [_product_brief(p) for p in merged],
                    }
                )
                seen_pairs.add((base, other))

    return Response({"count": len(duplicates), "groups": duplicates})


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
def merge_products(request):
    tenant = get_tenant_for_request(request)
    check_entitlement(tenant, "anti_duplicates")
    master_id = request.data.get("master_id")
    merge_ids = request.data.get("merge_ids") or []
    if not master_id or not merge_ids:
        return Response({"detail": "master_id et merge_ids requis."}, status=400)

    master = Product.all_objects.filter(id=master_id, tenant=tenant).select_related("service").first()
    if not master:
        return Response({"detail": "Produit maître introuvable."}, status=404)

    others = list(
        Product.all_objects.filter(id__in=merge_ids, tenant=tenant, service=master.service).exclude(id=master.id)
    )
    if not others:
        return Response({"detail": "Aucun doublon valide à fusionner."}, status=400)

    months = {master.inventory_month, *[p.inventory_month for p in others]}
    if len(months) > 1:
        return Response({"detail": "Fusion limitée à un même mois d’inventaire."}, status=400)

    with transaction.atomic():
        summary = {
            "before": _product_brief(master),
            "merged_ids": [p.id for p in others],
        }

        master.quantity = sum(float(p.quantity or 0) for p in [master, *others])
        for field in [
            "category",
            "barcode",
            "internal_sku",
            "variant_name",
            "variant_value",
            "purchase_price",
            "selling_price",
            "tva",
            "dlc",
            "lot_number",
            "min_qty",
            "conversion_unit",
            "conversion_factor",
            "brand",
            "supplier",
            "notes",
        ]:
            current = getattr(master, field, None)
            if current:
                continue
            for p in others:
                value = getattr(p, field, None)
                if value:
                    setattr(master, field, value)
                    break

        master.save()

        LossEvent.objects.filter(product_id__in=[p.id for p in others]).update(product=master)

        for dup in others:
            dup.is_archived = True
            dup.archived_at = timezone.now()
            dup.save(update_fields=["is_archived", "archived_at"])
            ProductMergeLog.objects.create(
                tenant=tenant,
                service=master.service,
                master_product=master,
                merged_product=dup,
                created_by=request.user if request.user.is_authenticated else None,
                summary=summary,
            )

        summary["after"] = _product_brief(master)

    return Response({"detail": "Fusion effectuée.", "master_id": master.id, "summary": summary})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ProductPermission])
def rituals(request):
    tenant = get_tenant_for_request(request)
    check_entitlement(tenant, "rituals")
    service = get_service_from_request(request)
    month = request.query_params.get("month")

    qs = Product.objects.filter(tenant=tenant, service=service)
    if month:
        qs = qs.filter(inventory_month=month)
    qs = _apply_retention(qs, tenant)

    losses_qs = LossEvent.objects.filter(tenant=tenant, service=service)
    if month:
        losses_qs = losses_qs.filter(inventory_month=month)
    losses_qs = _apply_retention(losses_qs, tenant)

    total_products = qs.count()
    low_stock = qs.filter(min_qty__isnull=False, quantity__lte=F("min_qty")).count()
    losses_count = losses_qs.count()
    missing_identifiers = qs.filter(
        (Q(barcode__isnull=True) | Q(barcode="")) & (Q(internal_sku__isnull=True) | Q(internal_sku=""))
    ).count()
    missing_prices = qs.filter(purchase_price__isnull=True, selling_price__isnull=True).count()
    missing_categories = qs.filter(Q(category__isnull=True) | Q(category="")).count()
    missing_suppliers = qs.filter(Q(supplier__isnull=True) | Q(supplier="")).count()
    missing_units = qs.filter(Q(unit__isnull=True) | Q(unit="")).count()

    now = timezone.now().date()
    dlc_30 = qs.filter(dlc__isnull=False, dlc__lte=now + timedelta(days=30)).count()
    dlc_90 = qs.filter(dlc__isnull=False, dlc__lte=now + timedelta(days=90)).count()

    service_type = getattr(service, "service_type", "other")

    missing_lots = qs.filter(lot_number__isnull=True).count()

    top_low_stock = list(
        qs.filter(min_qty__isnull=False, quantity__lte=F("min_qty"))
        .order_by("quantity")
        .values_list("name", flat=True)[:3]
    )
    top_losses = list(
        losses_qs.values("product__name")
        .annotate(total_qty=Sum("quantity"))
        .order_by("-total_qty")[:3]
    )

    def add_priority(acc, label, value, tone, hint):
        if value:
            acc.append({"label": label, "value": value, "tone": tone, "hint": hint})

    insights = []
    if total_products == 0:
        insights.append("Aucun produit suivi ce mois-ci : importez un catalogue pour démarrer le rituel.")
    if missing_identifiers:
        insights.append(f"{missing_identifiers} produit(s) sans identifiant : risque élevé de doublons.")
    if missing_prices:
        insights.append(f"{missing_prices} produit(s) sans prix : marge et pertes difficiles à mesurer.")
    if missing_categories:
        insights.append(f"{missing_categories} produit(s) sans catégorie : pilotage moins précis.")
    if missing_suppliers:
        insights.append(f"{missing_suppliers} produit(s) sans fournisseur : suivi d’achats incomplet.")
    if missing_units:
        insights.append(f"{missing_units} produit(s) sans unité claire : risques d’erreurs de comptage.")
    if dlc_30:
        insights.append(f"{dlc_30} produit(s) à moins de 30 jours d’échéance.")
    if low_stock:
        insights.append(f"{low_stock} produit(s) sous stock minimum.")
    if losses_count:
        insights.append(f"{losses_count} perte(s) enregistrée(s) ce mois.")
    if top_low_stock:
        insights.append("Stocks bas prioritaires : " + ", ".join(top_low_stock[:3]) + ".")
    if top_losses:
        focus = ", ".join(
            f"{item['product__name']} ({int(item['total_qty'] or 0)}u)"
            for item in top_losses
            if item.get("product__name")
        )
        if focus:
            insights.append(f"Pertes concentrées : {focus}.")

    score = 100
    score -= min(missing_identifiers * 2, 18)
    score -= min(missing_prices * 2, 16)
    score -= min(missing_categories, 10)
    score -= min(missing_suppliers, 10)
    score -= min(missing_units, 8)
    score -= min(dlc_30 * 3, 24)
    score -= min(low_stock * 2, 20)
    score -= min(losses_count * 3, 20)
    score = max(40, score)
    if score >= 80:
        status = "Stock sain"
    elif score >= 60:
        status = "À surveiller"
    else:
        status = "Priorités urgentes"

    dynamic_actions = []
    if missing_identifiers:
        dynamic_actions.append({"label": "Activer SKU / code-barres", "href": "/app/settings"})
    if missing_prices:
        dynamic_actions.append({"label": "Compléter les prix manquants", "href": "/app/products"})
    if missing_suppliers:
        dynamic_actions.append({"label": "Renseigner les fournisseurs clés", "href": "/app/products"})
    if dlc_30:
        dynamic_actions.append({"label": "Contrôler les DLC proches", "href": "/app/inventory"})
    if low_stock:
        dynamic_actions.append({"label": "Ajuster les stocks bas", "href": "/app/inventory"})
    if losses_count:
        dynamic_actions.append({"label": "Analyser les pertes", "href": "/app/losses"})

    ritual_map = {
        "grocery_food": {
            "title": "Rituel épicerie",
            "items": [
                {"label": "DLC < 30j", "value": dlc_30},
                {"label": "Stocks bas", "value": low_stock},
                {"label": "Pertes enregistrées", "value": losses_count},
            ],
            "priorities": [
                *([] if dlc_30 == 0 else [{"label": "DLC à moins de 30j", "value": dlc_30, "tone": "danger", "hint": "Rotation rapide"}]),
                *([] if low_stock == 0 else [{"label": "Stocks sous seuil", "value": low_stock, "tone": "warning", "hint": "Risque de rupture"}]),
                *([] if losses_count == 0 else [{"label": "Pertes du mois", "value": losses_count, "tone": "info", "hint": "À expliquer"}]),
                *([] if missing_identifiers == 0 else [{"label": "Produits sans identifiant", "value": missing_identifiers, "tone": "warning", "hint": "Doublons probables"}]),
            ],
            "checklist": [
                {"label": "Contrôler les DLC < 30j", "href": "/app/inventory"},
                {"label": "Ajuster les stocks bas", "href": "/app/inventory"},
                {"label": "Analyser les pertes", "href": "/app/losses"},
            ],
            "actions": [
                {"label": "Vérifier les stocks", "href": "/app/inventory"},
                {"label": "Analyser les pertes", "href": "/app/losses"},
            ],
        },
        "bakery": {
            "title": "Rituel boulangerie",
            "items": [
                {"label": "Pertes matières premières", "value": losses_count},
                {"label": "DLC < 30j", "value": dlc_30},
                {"label": "Stocks bas", "value": low_stock},
            ],
            "priorities": [
                *([] if losses_count == 0 else [{"label": "Pertes matières premières", "value": losses_count, "tone": "danger", "hint": "Causes ciblées"}]),
                *([] if low_stock == 0 else [{"label": "Stocks bas", "value": low_stock, "tone": "warning", "hint": "Risque de rupture"}]),
                *([] if dlc_30 == 0 else [{"label": "DLC à moins de 30j", "value": dlc_30, "tone": "info", "hint": "Rotation"}]),
                *([] if missing_prices == 0 else [{"label": "Prix d’achat manquants", "value": missing_prices, "tone": "info", "hint": "Impact pertes"}]),
            ],
            "checklist": [
                {"label": "Revoir les pertes sur matières premières", "href": "/app/losses"},
                {"label": "Planifier les achats pour les stocks bas", "href": "/app/inventory"},
                {"label": "Contrôler les DLC à venir", "href": "/app/inventory"},
            ],
            "actions": [
                {"label": "Voir les pertes", "href": "/app/losses"},
                {"label": "Ajuster les stocks", "href": "/app/inventory"},
            ],
        },
        "pharmacy_parapharmacy": {
            "title": "Rituel pharmacie",
            "items": [
                {"label": "DLC < 30j", "value": dlc_30},
                {"label": "DLC < 90j", "value": dlc_90},
                {"label": "Lots à vérifier", "value": missing_lots},
            ],
            "priorities": [
                *([] if dlc_30 == 0 else [{"label": "DLC à moins de 30j", "value": dlc_30, "tone": "danger", "hint": "Contrôle immédiat"}]),
                *([] if dlc_90 == 0 else [{"label": "DLC à 90j", "value": dlc_90, "tone": "warning", "hint": "Anticiper"}]),
                *([] if missing_lots == 0 else [{"label": "Lots manquants", "value": missing_lots, "tone": "info", "hint": "Traçabilité"}]),
                *([] if missing_suppliers == 0 else [{"label": "Fournisseurs manquants", "value": missing_suppliers, "tone": "info", "hint": "Suivi achats"}]),
            ],
            "checklist": [
                {"label": "Contrôler les dates proches", "href": "/app/inventory"},
                {"label": "Compléter les lots manquants", "href": "/app/products"},
                {"label": "Revoir les stocks sensibles", "href": "/app/inventory"},
            ],
            "actions": [
                {"label": "Revoir les fiches produits", "href": "/app/products"},
                {"label": "Contrôler les dates", "href": "/app/inventory"},
            ],
        },
    }
    ritual = ritual_map.get(service_type, ritual_map["grocery_food"]).copy()

    rituals_payload = [
        {
            "id": service_type,
            "title": ritual["title"],
            "summary": f"{total_products} produit(s) suivis · {status} ({score}/100).",
            "items": ritual["items"],
            "priorities": ritual.get("priorities", []),
            "checklist": ritual.get("checklist", []),
            "insights": insights[:3],
            "actions": (dynamic_actions + ritual["actions"])[:4],
            "status": status,
            "score": score,
        }
    ]

    return Response({"month": month, "rituals": rituals_payload})

def _env_int(name: str, default: int) -> int:
    try:
        v = int(str(os.getenv(name, "")).strip() or default)
        return v if v > 0 else default
    except Exception:
        return default


def _env_str(name: str, default: str) -> str:
    return (os.getenv(name, "") or default).strip()


def _sha256_hex(data: bytes) -> str:
    try:
        return hashlib.sha256(data or b"").hexdigest()
    except Exception:
        return ""


def _ocrspace_extract_text_from_pdf_bytes(pdf_bytes: bytes, *, language="fre", timeout_seconds=25) -> str:
    """
    OCR.Space - envoie le PDF et récupère le texte OCR.
    Retourne "" si échec.
    """
    api_key = _env_str("OCR_SPACE_API_KEY", "")
    if not api_key or not pdf_bytes:
        return ""

    try:
        resp = requests.post(
            "https://api.ocr.space/parse/image",
            files={"file": ("document.pdf", pdf_bytes, "application/pdf")},
            data={
                "apikey": api_key,
                "language": language,          # "fre" ou "eng" ou "fre,eng" (OCR.Space)
                "isOverlayRequired": "false",
                "OCREngine": "2",
                # On force un parsing “texte” simple
                "detectOrientation": "true",
                "scale": "true",
            },
            timeout=timeout_seconds,
        )

        # OCR.Space peut renvoyer 200 même en erreur => inspect JSON
        try:
            data = resp.json()
        except Exception:
            return ""

        if not isinstance(data, dict):
            return ""

        if data.get("IsErroredOnProcessing"):
            # Ex: quota / file not supported
            return ""

        parsed = data.get("ParsedResults") or []
        if not parsed:
            return ""

        # Concat pages
        text_parts = []
        for p in parsed:
            t = (p.get("ParsedText") or "").strip()
            if t:
                text_parts.append(t)

        return "\n".join(text_parts).strip()

    except requests.Timeout:
        return ""
    except Exception:
        logger.exception("OCR_SPACE_FAILED")
        return ""


def _ocr_receipt_text(pdf_bytes: bytes) -> str:
    """
    Wrapper avec cache pour éviter de repayer plusieurs fois le même PDF.
    """
    provider = _env_str("RECEIPTS_OCR_PROVIDER", RECEIPTS_OCR_PROVIDER_DEFAULT).lower()
    timeout_seconds = _env_int("RECEIPTS_OCR_TIMEOUT_SECONDS", RECEIPTS_OCR_TIMEOUT_DEFAULT)
    ttl = _env_int("RECEIPTS_OCR_CACHE_TTL_SECONDS", RECEIPTS_OCR_CACHE_TTL_DEFAULT)

    # Cache key coûts
    h = _sha256_hex(pdf_bytes)[:16]
    cache_key = f"receipts:ocr:{provider}:{h}"
    try:
        cached = cache.get(cache_key)
        if isinstance(cached, str) and cached.strip():
            return cached.strip()
    except Exception:
        cached = None

    if provider == "ocrspace":
        txt = _ocrspace_extract_text_from_pdf_bytes(pdf_bytes, language="fre", timeout_seconds=timeout_seconds)
    else:
        # provider inconnu -> pas d’OCR
        txt = ""

    txt = _normalize_receipt_text(txt).strip()

    try:
        if txt:
            cache.set(cache_key, txt, ttl)
    except Exception:
        pass

    return txt
def _parse_receipt_rows(file_obj, file_name):
    file_name = file_name or "import"
    if file_name.lower().endswith(".csv"):
        raw = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(raw))
        rows = list(reader)
        meta = _extract_invoice_meta_from_rows(rows)
        return rows, "csv", meta

    if file_name.lower().endswith(".pdf"):
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise exceptions.ValidationError("PDF non supporté sur ce déploiement.") from exc

        # Lire une seule fois pour:
        # - pypdf via BytesIO
        # - OCR fallback si besoin
        pdf_bytes = file_obj.read()
        try:
            file_obj.seek(0)
        except Exception:
            pass

        max_pages = _env_int("RECEIPTS_OCR_MAX_PAGES", RECEIPTS_OCR_MAX_PAGES_DEFAULT)

        reader = PdfReader(io.BytesIO(pdf_bytes))
        page_texts = []
        for idx, page in enumerate(reader.pages):
            if max_pages and idx >= max_pages:
                break
            text = ""
            try:
                text = page.extract_text(extraction_mode="layout") or ""
            except TypeError:
                text = page.extract_text() or ""
            except Exception:
                text = ""
            if not text:
                try:
                    text = page.extract_text() or ""
                except Exception:
                    text = ""
            if text:
                page_texts.append(text)

        raw_text = "\n".join(page_texts)
        raw_text = _normalize_receipt_text(raw_text).strip()

        # ✅ Fallback OCR si PDF scanné / vide
        if not raw_text or len(raw_text) < RECEIPTS_OCR_MIN_TEXT_LEN:
            ocr_text = _ocr_receipt_text(pdf_bytes)
            if ocr_text and len(ocr_text) >= RECEIPTS_OCR_MIN_TEXT_LEN:
                raw_text = ocr_text
                logger.info("receipts_ocr_used", extra={"file": file_name, "pages": max_pages})

        raw_text = _normalize_receipt_text(raw_text)
        meta = _extract_invoice_meta_from_text(raw_text)

        lines = [_normalize_receipt_text(line).strip() for line in raw_text.splitlines() if line.strip()]
        lines = _merge_receipt_lines(lines)
        if not lines:
            return [], "pdf", meta

        header = lines[0].lower()
        header_idx = None
        for idx, line in enumerate(lines[:20]):
            lower = line.lower()
            if any(k in lower for k in ("produit", "designation", "libelle", "article", "quantite", "qte")) and (
                ";" in line or "," in line or "\t" in line or "|" in line
            ):
                header_idx = idx
                header = lower
                break

        has_delim = ";" in header or "," in header or "\t" in header or "|" in header
        if has_delim and any(k in header for k in ("produit", "designation", "libelle", "qty", "quantite", "qte", "prix")):
            delimiter = ";" if ";" in header else "," if "," in header else "\t" if "\t" in header else "|"
            csv_start = header_idx if header_idx is not None else 0
            reader = csv.DictReader(io.StringIO("\n".join(lines[csv_start:])), delimiter=delimiter)
            rows = list(reader)
            if rows:
                meta_from_rows = _extract_invoice_meta_from_rows(rows)
                return rows, "pdf", meta_from_rows or meta

        rows = []
        for line in lines:
            parsed = _parse_receipt_line(line)
            if parsed:
                rows.append(parsed)
        return rows, "pdf", meta

    raise exceptions.ValidationError("Format non supporté (CSV ou PDF).")


def _normalize_receipt_text(value):
    if not value:
        return ""
    return (
        str(value)
        .replace("\u00a0", " ")
        .replace("\u202f", " ")
        .replace("\u2007", " ")
    )


def _merge_receipt_lines(lines):
    if not lines:
        return []

    def has_price(val):
        return bool(re.search(r"\\d+[.,]\\d{2}", val))

    def has_words(val):
        return bool(re.search(r"[A-Za-zÀ-ÿ]", val))

    def is_name_only(val):
        return has_words(val) and not has_price(val)

    merged = []
    idx = 0
    while idx < len(lines):
        current = _normalize_receipt_text(lines[idx]).strip()
        if not current:
            idx += 1
            continue
        next_line = _normalize_receipt_text(lines[idx + 1]).strip() if idx + 1 < len(lines) else ""
        if is_name_only(current) and next_line and has_price(next_line):
            merged.append(f"{current} {next_line}")
            idx += 2
            continue
        merged.append(current)
        idx += 1
    return merged


RECEIPT_IGNORE_TOKENS = (
    "facture",
    "total",
    "ht",
    "ttc",
    "montant",
    "reglement",
    "iban",
    "date",
    "client",
    "adresse",
    "num",
    "numero",
    "référence",
    "reference",
)
RECEIPT_NAME_BLACKLIST = {
    "facture",
    "invoice",
    "total",
    "ttc",
    "ht",
    "tva",
    "montant",
    "date",
    "client",
    "adresse",
    "reglement",
    "iban",
}
RECEIPT_UNITS = {"kg", "g", "l", "ml", "pcs", "pc", "un", "u", "unite", "unité"}
RECEIPT_INVOICE_PATTERNS = (
    r"(?:facture|invoice|bon)\\s*(?:n[o°º]?|num(?:ero)?|#)?\\s*[:\\-]?\\s*([A-Z0-9][A-Z0-9\\-_/]{2,})",
    r"\\bF\\d{4,}\\b",
    r"\\b[A-Z]{2,5}[-/]\\d{3,}\\b",
)
RECEIPT_DATE_PATTERNS = (
    r"\\b\\d{4}-\\d{2}-\\d{2}\\b",
    r"\\b\\d{1,2}[\\-/]\\d{1,2}[\\-/]\\d{2,4}\\b",
)


def _normalize_invoice_number(value):
    if not value:
        return ""
    raw = re.sub(r"\\s+", "", str(value).strip())
    return raw.upper()


def _parse_invoice_date(value):
    if not value:
        return None
    raw = str(value).strip().replace(".", "/")
    if re.match(r"^\\d{4}-\\d{2}-\\d{2}$", raw):
        return parse_date(raw)
    match = re.match(r"^(\\d{1,2})[\\-/](\\d{1,2})[\\-/](\\d{2,4})$", raw)
    if not match:
        return None
    day, month, year = match.groups()
    if len(year) == 2:
        year = f"20{year}"
    return parse_date(f"{year}-{int(month):02d}-{int(day):02d}")


def _extract_invoice_meta_from_text(text):
    meta = {"invoice_number": "", "invoice_date": None}
    if not text:
        return meta
    for pattern in RECEIPT_INVOICE_PATTERNS:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            meta["invoice_number"] = _normalize_invoice_number(
                match.group(1) if match.groups() else match.group(0)
            )
            break
    for pattern in RECEIPT_DATE_PATTERNS:
        match = re.search(pattern, text)
        if match:
            meta["invoice_date"] = _parse_invoice_date(match.group(0))
            if meta["invoice_date"]:
                break
    return meta


def _extract_invoice_meta_from_rows(rows):
    meta = {"invoice_number": "", "invoice_date": None}
    for row in rows or []:
        lower = {str(k).strip().lower(): v for k, v in (row or {}).items()}
        if not meta["invoice_number"]:
            for key in ("invoice_number", "facture", "numero_facture", "num_facture", "invoice"):
                value = lower.get(key)
                if value:
                    meta["invoice_number"] = _normalize_invoice_number(value)
                    break
        if not meta["invoice_date"]:
            for key in ("invoice_date", "date_facture", "date"):
                parsed = _parse_invoice_date(lower.get(key))
                if parsed:
                    meta["invoice_date"] = parsed
                    break
        if meta["invoice_number"] and meta["invoice_date"]:
            break
    return meta


def _parse_receipt_line(line: str):
    if not line:
        return None
    line = _normalize_receipt_text(line).strip()
    lower = line.lower()
    if any(lower.startswith(token) for token in RECEIPT_IGNORE_TOKENS):
        return None

    def to_number(token):
        cleaned = re.sub(r"[^0-9,\\.]", "", token)
        if not cleaned:
            return None
        cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None

    def parse_columns(columns):
        if not columns:
            return None
        cols = [c.strip() for c in columns if c.strip()]
        if len(cols) < 2:
            return None
        barcode = ""
        for idx, col in enumerate(list(cols)):
            compact = col.replace(" ", "")
            if re.fullmatch(r"\\d{8,14}", compact):
                barcode = compact
                cols.pop(idx)
                break

        price_candidates = []
        for idx in range(len(cols) - 1, -1, -1):
            num = to_number(cols[idx])
            if num is not None:
                price_candidates.append((idx, num))
        if not price_candidates:
            qty_idx = None
            qty_val = None
            unit = ""
            for idx in range(len(cols) - 1, -1, -1):
                part = cols[idx]
                match = re.fullmatch(r"(\\d+[.,]?\\d*)\\s*([a-zA-Z]{1,4})?", part)
                if match:
                    qty_val = to_number(match.group(1))
                    unit = (match.group(2) or "").lower()
                    if unit and unit not in RECEIPT_UNITS:
                        unit = ""
                    qty_idx = idx
                    break
            if qty_idx is not None and not unit and qty_idx + 1 < len(cols):
                unit_candidate = cols[qty_idx + 1].strip().lower()
                if unit_candidate in RECEIPT_UNITS:
                    unit = unit_candidate
            if qty_val is None:
                return None
            name_parts = []
            for idx, col in enumerate(cols):
                if idx == qty_idx:
                    continue
                if col.lower().strip(".,;") in RECEIPT_UNITS:
                    continue
                if re.fullmatch(r"\\d+[.,]?\\d*", col):
                    continue
                name_parts.append(col)
            name = " ".join(name_parts).strip()
            if not name:
                return None
            tokens = _tokenize_name(name)
            if not tokens:
                return None
            if len(tokens) == 1 and (tokens[0] in RECEIPT_NAME_BLACKLIST or len(tokens[0]) < 3):
                return None
            return {
                "name": name,
                "quantity": qty_val if qty_val is not None else "",
                "unit": unit,
                "purchase_price": "",
                "tva": "",
                "category": "",
                "barcode": barcode,
                "internal_sku": "",
            }
        price_candidates = list(reversed(price_candidates))
        price_idx, price_val = price_candidates[-1]

        qty_idx = None
        qty_val = None
        unit = ""
        for idx in range(price_idx - 1, -1, -1):
            part = cols[idx]
            match = re.fullmatch(r"(\\d+[.,]?\\d*)\\s*([a-zA-Z]{1,4})?", part)
            if match:
                qty_val = to_number(match.group(1))
                unit = (match.group(2) or "").lower()
                if unit and unit not in RECEIPT_UNITS:
                    unit = ""
                qty_idx = idx
                break
        if qty_idx is not None and not unit and qty_idx + 1 < len(cols):
            unit_candidate = cols[qty_idx + 1].strip().lower()
            if unit_candidate in RECEIPT_UNITS:
                unit = unit_candidate

        if qty_val is not None and len(price_candidates) >= 2:
            total_price = price_candidates[-1][1]
            for idx, num in reversed(price_candidates[:-1]):
                if qty_idx is not None and idx == qty_idx:
                    continue
                if num <= 0:
                    continue
                if abs((num * qty_val) - total_price) < 0.05:
                    price_idx, price_val = idx, num
                    break

        tva = None
        for col in cols:
            if "%" in col:
                rate = to_number(col)
                if rate is not None and 0 <= rate <= 30:
                    tva = rate
                    break

        name_parts = []
        for idx, col in enumerate(cols):
            if idx == price_idx or (qty_idx is not None and idx == qty_idx):
                continue
            if col.lower().strip(".,;") in RECEIPT_UNITS:
                continue
            if re.fullmatch(r"\\d+[.,]?\\d*", col):
                continue
            name_parts.append(col)
        name = " ".join(name_parts).strip()
        if not name:
            return None
        tokens = _tokenize_name(name)
        if not tokens:
            return None
        if len(tokens) == 1 and (tokens[0] in RECEIPT_NAME_BLACKLIST or len(tokens[0]) < 3):
            return None
        return {
            "name": name,
            "quantity": qty_val if qty_val is not None else "",
            "unit": unit,
            "purchase_price": price_val,
            "tva": tva if tva is not None else "",
            "category": "",
            "barcode": barcode,
            "internal_sku": "",
        }

    for delim in (";", "|", "\t"):
        if delim in line:
            parsed = parse_columns(line.split(delim))
            if parsed:
                return parsed

    columns = [c.strip() for c in re.split(r"\\s{2,}|\\t|\\|", line) if c.strip()]
    parsed_columns = parse_columns(columns)
    if parsed_columns:
        return parsed_columns

    def parse_regex(raw_line):
        price_hits = re.findall(r"(\\d+[.,]\\d{2})", raw_line)
        if not price_hits:
            return None
        price_raw = price_hits[-1]
        price_val = to_number(price_raw)
        if price_val is None:
            return None
        prefix = raw_line[: raw_line.rfind(price_raw)].strip()
        prefix = re.sub(r"[€$£]\\s*$", "", prefix).strip()
        prefix = re.sub(r"[×x]\\s*$", "", prefix).strip()

        qty_val = None
        unit = ""
        qty_match = re.search(r"(\\d+[.,]?\\d*)\\s*(kg|g|l|ml|pcs|pc|u|un|unite|unité)?\\s*$", prefix, flags=re.I)
        barcode = ""
        barcode_match = re.search(r"\\b(\\d{8,14})\\b", prefix)
        if barcode_match:
            barcode = barcode_match.group(1)
            prefix = prefix.replace(barcode, " ").strip()

        if qty_match:
            qty_val = to_number(qty_match.group(1))
            unit = (qty_match.group(2) or "").lower()
            if unit and unit not in RECEIPT_UNITS:
                unit = ""
            name = prefix[: qty_match.start()].strip()
        else:
            name = prefix

        if not name:
            return None
        tokens = _tokenize_name(name)
        if not tokens:
            return None
        if len(tokens) == 1 and (tokens[0] in RECEIPT_NAME_BLACKLIST or len(tokens[0]) < 3):
            return None

        tva = None
        for token in re.split(r"\\s+", raw_line):
            if "%" in token:
                rate = to_number(token)
                if rate is not None and 0 <= rate <= 30:
                    tva = rate
                    break

        return {
            "name": name,
            "quantity": qty_val if qty_val is not None else "",
            "unit": unit,
            "purchase_price": price_val,
            "tva": tva if tva is not None else "",
            "category": "",
            "barcode": barcode,
            "internal_sku": "",
        }

    tokens = [t for t in re.split(r"\\s+", line) if t]
    if len(tokens) < 2:
        parsed_regex = parse_regex(line)
        return parsed_regex

    barcode = ""
    for t in tokens:
        if re.fullmatch(r"\\d{8,14}", t):
            barcode = t
            break

    number_positions = []
    for idx, t in enumerate(tokens):
        num = to_number(t)
        if num is not None:
            number_positions.append((idx, num))

    if not number_positions:
        return None

    price_idx, price_val = number_positions[-1]
    qty_val = None
    if len(number_positions) >= 2:
        qty_idx, qty_val = number_positions[-2]
        if qty_val > 10000:
            qty_val = None
    else:
        qty_idx = None

    unit = ""
    for idx, t in enumerate(tokens):
        t_clean = t.lower().strip(".,;")
        if t_clean in RECEIPT_UNITS:
            unit = t_clean
            break
        token_match = re.fullmatch(r"(\\d+[.,]?\\d*)(kg|g|l|ml|pcs|pc|u|un|unite|unité)", t_clean)
        if token_match and qty_idx is None:
            qty_val = to_number(token_match.group(1))
            unit = token_match.group(2).lower()
            qty_idx = idx

    if len(number_positions) == 1 and unit:
        qty_idx = price_idx
        qty_val = price_val
        price_val = ""
        price_idx = None

    tva = None
    for t in tokens:
        if "%" in t:
            rate = to_number(t)
            if rate is not None and 0 <= rate <= 30:
                tva = rate
                break

    name_tokens = []
    for idx, t in enumerate(tokens):
        if barcode and t == barcode:
            continue
        if (price_idx is not None and idx == price_idx) or (qty_idx is not None and idx == qty_idx):
            continue
        if t.lower().strip(".,;") in RECEIPT_UNITS:
            continue
        if re.fullmatch(r"\\d+[.,]?\\d*", t):
            continue
        name_tokens.append(t)

    name = " ".join(name_tokens).strip()
    if not name:
        parsed_regex = parse_regex(line)
        return parsed_regex
    tokens = _tokenize_name(name)
    if not tokens:
        parsed_regex = parse_regex(line)
        return parsed_regex
    if len(tokens) == 1 and (tokens[0] in RECEIPT_NAME_BLACKLIST or len(tokens[0]) < 3):
        parsed_regex = parse_regex(line)
        return parsed_regex

    return {
        "name": name,
        "quantity": qty_val if qty_val is not None else "",
        "unit": unit,
        "purchase_price": price_val,
        "tva": tva if tva is not None else "",
        "category": "",
        "barcode": barcode,
        "internal_sku": "",
    }


def _normalize_receipt_row(row):
    def pick(*keys):
        for key in keys:
            value = row.get(key)
            if value not in (None, ""):
                return value
        return ""

    return {
        "name": pick("name", "product", "designation", "libelle", "article"),
        "quantity": pick("quantity", "qty", "quantite", "qte"),
        "unit": pick("unit", "uom", "unite"),
        "purchase_price": pick("purchase_price", "prix_achat", "price", "prix"),
        "tva": pick("tva", "vat", "tax", "taux_tva"),
        "category": pick("category", "categorie", "famille", "rayon"),
        "barcode": pick("barcode", "ean", "code_barres"),
        "internal_sku": pick("internal_sku", "sku", "ref", "reference"),
    }


def _compute_receipt_hash(rows, supplier_name, invoice_number, invoice_date, received_at):
    if not rows:
        return ""
    normalized = []
    for row in rows:
        normalized.append(
            {
                "name": (row.get("name") or "").strip().lower(),
                "quantity": str(row.get("quantity") or "").strip(),
                "unit": (row.get("unit") or "").strip().lower(),
                "purchase_price": str(row.get("purchase_price") or "").strip(),
                "tva": str(row.get("tva") or "").strip(),
                "barcode": str(row.get("barcode") or "").strip(),
                "internal_sku": str(row.get("internal_sku") or "").strip(),
            }
        )
    payload = {
        "supplier": (supplier_name or "").strip().lower(),
        "invoice_number": (invoice_number or "").strip().lower(),
        "invoice_date": str(invoice_date or ""),
        "received_at": str(received_at or ""),
        "lines": normalized,
    }
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _match_product_from_line(tenant, service, line, month):
    code = (line.get("barcode") or "").strip()
    sku = (line.get("internal_sku") or "").strip()
    name = (line.get("name") or "").strip()

    if code:
        found = Product.objects.filter(tenant=tenant, service=service, inventory_month=month, barcode=code).first()
        if found:
            return found
    if sku:
        found = Product.objects.filter(tenant=tenant, service=service, inventory_month=month, internal_sku=sku).first()
        if found:
            return found
    if name:
        return Product.objects.filter(tenant=tenant, service=service, inventory_month=month, name__iexact=name).first()
    return None


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
def import_receipt(request):
    tenant = get_tenant_for_request(request)
    check_entitlement(tenant, "receipts_import")
    _enforce_receipts_import_quota(tenant)

    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response({"detail": "Fichier requis."}, status=400)

    service = get_service_from_request(request)
    supplier_name = (request.data.get("supplier_name") or "").strip()
    received_at_raw = request.data.get("received_at") or ""
    received_at = _parse_invoice_date(received_at_raw) or parse_date(received_at_raw) or timezone.now().date()

    rows, source, meta = _parse_receipt_rows(file_obj, file_obj.name)
    if len(rows) > RECEIPT_IMPORT_MAX_LINES:
        rows = rows[:RECEIPT_IMPORT_MAX_LINES]
    raw_rows_count = len(rows)

    invoice_number = _normalize_invoice_number(
        request.data.get("invoice_number") or (meta or {}).get("invoice_number") or ""
    )
    invoice_date = _parse_invoice_date(request.data.get("invoice_date") or "")
    if not invoice_date:
        invoice_date = (meta or {}).get("invoice_date")

    if invoice_number:
        exists = Receipt.objects.filter(tenant=tenant, invoice_number__iexact=invoice_number).exists()
        if exists:
            return Response(
                {
                    "detail": "Facture déjà importée pour ce commerce.",
                    "code": "RECEIPT_DUPLICATE_INVOICE",
                },
                status=409,
            )

    cleaned_rows = []
    for row in rows:
        cleaned = _normalize_receipt_row(row)
        if not cleaned["name"]:
            continue
        cleaned_rows.append(cleaned)

    if not cleaned_rows:
        return Response(
            {
                "detail": "Aucune ligne exploitable n’a été trouvée dans ce fichier. Vérifiez le format.",
                "code": "RECEIPT_EMPTY",
            },
            status=422,
        )

    import_hash = _compute_receipt_hash(
        cleaned_rows,
        supplier_name,
        invoice_number,
        invoice_date,
        received_at,
    )
    if not invoice_number and import_hash:
        exists = Receipt.objects.filter(tenant=tenant, import_hash=import_hash).exists()
        if exists:
            return Response(
                {
                    "detail": "Cette facture a déjà été importée pour ce commerce.",
                    "code": "RECEIPT_DUPLICATE_INVOICE",
                },
                status=409,
            )

    supplier = None
    if supplier_name:
        supplier = Supplier.objects.filter(tenant=tenant, name__iexact=supplier_name).first()
        if not supplier:
            for candidate in Supplier.objects.filter(tenant=tenant):
                if supplier_name in (candidate.aliases or []):
                    supplier = candidate
                    break
        if not supplier:
            supplier = Supplier.objects.create(tenant=tenant, name=supplier_name)
        elif supplier.name.strip().lower() != supplier_name.lower():
            supplier.add_alias(supplier_name)

    receipt = Receipt.objects.create(
        tenant=tenant,
        service=service,
        supplier=supplier,
        supplier_name=supplier_name,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        import_hash=import_hash,
        source=source,
        file_name=file_obj.name or "",
        received_at=received_at,
        created_by=request.user if request.user.is_authenticated else None,
    )

    month = timezone.now().strftime("%Y-%m")
    lines_payload = []
    skipped = max(0, raw_rows_count - len(cleaned_rows))
    for idx, cleaned in enumerate(cleaned_rows, start=1):
        try:
            qty = float(cleaned["quantity"]) if cleaned["quantity"] not in ("", None) else 1
        except (TypeError, ValueError):
            qty = 1
        if qty <= 0:
            qty = 1
        try:
            price = float(cleaned["purchase_price"]) if cleaned["purchase_price"] not in ("", None) else None
        except (TypeError, ValueError):
            price = None
        try:
            tva = float(cleaned["tva"]) if cleaned.get("tva") not in ("", None) else None
        except (TypeError, ValueError):
            tva = None

        matched = _match_product_from_line(tenant, service, cleaned, month)

        line = ReceiptLine.objects.create(
            receipt=receipt,
            line_number=idx,
            raw_name=cleaned["name"],
            quantity=qty,
            unit=cleaned["unit"] or "pcs",
            purchase_price=price,
            tva=tva,
            category=cleaned.get("category") or "",
            barcode=cleaned["barcode"] or "",
            internal_sku=cleaned["internal_sku"] or "",
            matched_product=matched,
            status="MATCHED" if matched else "PENDING",
        )
        lines_payload.append(
            {
                "id": line.id,
                "name": line.raw_name,
                "quantity": float(line.quantity or 0),
                "unit": line.unit,
                "purchase_price": line.purchase_price,
                "tva": line.tva,
                "category": line.category,
                "barcode": line.barcode,
                "internal_sku": line.internal_sku,
                "matched_product_id": matched.id if matched else None,
            }
        )

    _log_receipt_import_event(
        tenant=tenant,
        user=request.user,
        params={"lines": len(lines_payload), "supplier": supplier_name, "source": source, "skipped": skipped},
    )
    logger.info(
        "receipts_imported",
        extra={
            "tenant_id": tenant.id,
            "service_id": getattr(service, "id", None),
            "lines": len(lines_payload),
            "skipped": skipped,
            "supplier": supplier_name,
            "source": source,
        },
    )

    return Response(
        {
            "receipt_id": receipt.id,
            "supplier": supplier_name,
            "invoice_number": receipt.invoice_number,
            "invoice_date": receipt.invoice_date.isoformat() if receipt.invoice_date else None,
        "received_at": receipt.received_at.isoformat() if receipt.received_at else None,
            "lines": lines_payload,
            "skipped_lines": skipped,
        },
        status=201,
    )


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
def apply_receipt(request, receipt_id):
    tenant = get_tenant_for_request(request)
    receipt = Receipt.objects.filter(id=receipt_id, tenant=tenant).select_related("service").first()
    if not receipt:
        return Response({"detail": "Réception introuvable."}, status=404)

    decisions = request.data.get("decisions") or []
    decisions_map = {str(d.get("line_id")): d for d in decisions if d.get("line_id")}
    month = timezone.now().strftime("%Y-%m")

    applied = 0
    with transaction.atomic():
        for line in receipt.lines.select_for_update():
            decision = decisions_map.get(str(line.id), {}) if decisions_map else {}
            action = decision.get("action") or ("match" if line.matched_product else "create")

            if action == "ignore":
                line.status = "IGNORED"
                line.save(update_fields=["status"])
                continue

            product = None
            if action == "match":
                product_id = decision.get("product_id") or (line.matched_product_id if line.matched_product_id else None)
                if product_id:
                    product = Product.objects.filter(id=product_id, tenant=tenant, service=receipt.service, inventory_month=month).first()
                if not product:
                    product = _match_product_from_line(tenant, receipt.service, {
                        "barcode": line.barcode,
                        "internal_sku": line.internal_sku,
                        "name": line.raw_name,
                    }, month)

            if action == "create" or product is None:
                product = Product.objects.create(
                    tenant=tenant,
                    service=receipt.service,
                    name=line.raw_name,
                    inventory_month=month,
                    quantity=0,
                    unit=line.unit or "pcs",
                    barcode=line.barcode or "",
                    internal_sku=line.internal_sku or "",
                    purchase_price=line.purchase_price,
                )

            updates = []
            if line.barcode and not product.barcode:
                product.barcode = line.barcode
                updates.append("barcode")
            if line.internal_sku and not product.internal_sku:
                product.internal_sku = line.internal_sku
                updates.append("internal_sku")
            if line.category and not product.category:
                product.category = line.category
                updates.append("category")
            if line.unit and not product.unit:
                product.unit = line.unit
                updates.append("unit")
            if line.purchase_price is not None and product.purchase_price in (None, "", 0):
                product.purchase_price = line.purchase_price
                updates.append("purchase_price")
            if line.tva is not None and product.tva in (None, ""):
                product.tva = line.tva
                updates.append("tva")

            product.quantity = float(product.quantity or 0) + float(line.quantity or 0)
            updates.append("quantity")
            product.save(update_fields=updates)
            line.matched_product = product
            line.status = "MATCHED" if action == "match" else "CREATED"
            line.save(update_fields=["matched_product", "status"])
            applied += 1

        receipt.status = "APPLIED"
        receipt.save(update_fields=["status"])
    logger.info(
        "receipt_applied",
        extra={
            "tenant_id": tenant.id,
            "receipt_id": receipt.id,
            "service_id": getattr(receipt.service, "id", None),
            "applied": applied,
        },
    )

    return Response({"detail": "Réception appliquée.", "applied": applied})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
def receipts_history(request):
    tenant = get_tenant_for_request(request)
    check_entitlement(tenant, "receipts_import")

    service_param = request.query_params.get("service")
    query = (request.query_params.get("q") or "").strip()
    date_param = (request.query_params.get("date") or "").strip()
    limit = _parse_positive_int(request.query_params.get("limit"), 30) or 30

    qs = (
        Receipt.objects.filter(tenant=tenant)
        .select_related("service", "supplier")
        .annotate(lines_count=Count("lines"))
        .order_by("-received_at", "-created_at")
    )

    if service_param and service_param != "all":
        try:
            qs = qs.filter(service_id=int(service_param))
        except (ValueError, TypeError):
            pass

    if query:
        q_date = _parse_invoice_date(query) or parse_date(query)
        filters = (
            Q(invoice_number__icontains=query)
            | Q(supplier_name__icontains=query)
            | Q(file_name__icontains=query)
        )
        if q_date:
            filters |= Q(received_at=q_date) | Q(invoice_date=q_date)
        qs = qs.filter(filters)

    if date_param:
        parsed_date = _parse_invoice_date(date_param) or parse_date(date_param)
        if parsed_date:
            qs = qs.filter(Q(received_at=parsed_date) | Q(invoice_date=parsed_date))

    receipts = list(qs[:limit])
    payload = [
        {
            "id": r.id,
            "service_id": r.service_id,
            "service_name": r.service.name if r.service else None,
            "supplier": r.supplier_name or (r.supplier.name if r.supplier else ""),
            "invoice_number": r.invoice_number,
            "invoice_date": r.invoice_date.isoformat() if r.invoice_date else None,
            "received_at": r.received_at.isoformat() if r.received_at else None,
            "status": r.status,
            "source": r.source,
            "lines_count": getattr(r, "lines_count", 0),
        }
        for r in receipts
    ]

    return Response({"count": len(payload), "results": payload})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, ManagerPermission])
@renderer_classes([PDFRenderer])
def labels_pdf(request):
    tenant = get_tenant_for_request(request)
    check_entitlement(tenant, "labels_pdf")
    _enforce_labels_pdf_quota(tenant)

    service_param = request.query_params.get("service")
    company_name = (request.query_params.get("company_name") or "").strip() or tenant.name
    currency_code = getattr(tenant, "currency_code", "EUR")
    promo = _parse_promo(request.query_params.get("promo_type"), request.query_params.get("promo_value"))
    fields_raw = request.query_params.get("fields") or ""
    fields = [f.strip() for f in fields_raw.split(",") if f.strip() in LABEL_ALLOWED_FIELDS]
    if promo and "price" not in fields:
        fields.append("price")

    ids = request.query_params.get("ids") or ""
    ids_list = [int(i) for i in ids.split(",") if i.strip().isdigit()]
    if not ids_list:
        return Response({"detail": "Sélectionnez au moins un produit."}, status=400)

    count_map = _parse_label_counts(request.query_params.get("counts"))

    qs = Product.objects.filter(tenant=tenant, id__in=ids_list)
    if service_param and service_param != "all":
        try:
            qs = qs.filter(service_id=int(service_param))
        except (ValueError, TypeError):
            pass
    if ids_list:
        ordering = Case(
            *[When(id=pid, then=pos) for pos, pid in enumerate(ids_list)],
            output_field=IntegerField(),
        )
        qs = qs.annotate(_order=ordering).order_by("_order", "name")

    products = list(qs[:LABELS_PDF_MAX_PRODUCTS])
    if not products:
        return Response({"detail": "Aucun produit disponible."}, status=404)

    label_entries = []
    for product in products:
        count = count_map.get(product.id, 1)
        label_entries.extend([product] * count)

    if len(label_entries) > LABELS_PDF_MAX_PRODUCTS:
        label_entries = label_entries[:LABELS_PDF_MAX_PRODUCTS]

    for product in products:
        if product.barcode or product.internal_sku:
            continue
        try:
            product.internal_sku = generate_auto_sku(tenant, product.service)
            product.save(update_fields=["internal_sku"])
        except exceptions.ValidationError as exc:
            return Response(exc.detail, status=400)

    width, height = A4
    margin_x = 0.18 * cm
    margin_y = 0.18 * cm
    cols = 3
    rows = 8
    label_w = (width - (margin_x * 2)) / cols
    label_h = (height - (margin_y * 2)) / rows

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    c.setTitle("StockScan Labels")

    def wrap_label_text(text, max_width, font_name="Helvetica-Bold", font_size=9, max_lines=2):
        if not text:
            return []
        words = str(text).split()
        lines, current = [], ""
        for w in words:
            test = (current + " " + w).strip()
            if c.stringWidth(test, font_name, font_size) <= max_width:
                current = test
            else:
                if current:
                    lines.append(current)
                current = w
        if current:
            lines.append(current)
        return lines[:max_lines]

    def draw_label(x, y, product):
        code = (product.barcode or product.internal_sku or "").strip()
        if not code:
            return False

        # tighter padding -> less empty
        pad_x = 6
        pad_y = 6

        # Header (name)
        name = (product.name or "").strip().upper()
        c.setFillColorRGB(0.08, 0.1, 0.13)

        name_font = 9.2
        c.setFont("Helvetica-Bold", name_font)
        max_name_w = label_w - (pad_x * 2)

        name_lines = wrap_label_text(name, max_name_w, "Helvetica-Bold", name_font, max_lines=2)
        y_top = y + label_h - pad_y - 2
        for line in name_lines:
            c.drawString(x + pad_x, y_top, line)
            y_top -= 10.2

        # Price block (bigger, better centered)
        base_price = product.selling_price if product.selling_price is not None else product.purchase_price
        show_price = "price" in fields
        price = _format_price(base_price, currency_code) if show_price and base_price is not None else None

        promo_price = _apply_promo(base_price, promo) if promo else None
        promo_active = promo_price is not None and show_price and base_price is not None

        # Barcode: bigger + less wasted space
        barcode_height = max(30, label_h * 0.30)
        barcode_obj = code128.Code128(code, barHeight=barcode_height, barWidth=0.88)

        barcode_y = y + pad_y + 10
        barcode_x = x + (label_w - barcode_obj.width) / 2

        # Compute price area just above barcode
        price_y = barcode_y + barcode_height + 18
        # Ensure it doesn't overlap name area
        price_y = max(price_y, y_top - 6)

        if show_price:
            if promo_active:
                old_price = _format_price(base_price, currency_code)
                new_price = _format_price(promo_price, currency_code)

                badge = (
                    f"-{int(promo['value'])}%"
                    if promo["type"] == "percent"
                    else f"-{_format_price(promo['value'], currency_code)}"
                )
                c.setFont("Helvetica", 7.8)
                c.setFillColorRGB(0.25, 0.29, 0.34)
                c.drawString(x + pad_x, price_y + 12, old_price or "")
                old_w = c.stringWidth(old_price or "", "Helvetica", 7.8)
                c.setLineWidth(1)
                c.line(x + pad_x, price_y + 10, x + pad_x + old_w, price_y + 10)

                c.setFont("Helvetica-Bold", 20.5)
                c.setFillColorRGB(0.08, 0.1, 0.13)
                c.drawCentredString(x + label_w / 2, price_y, new_price or "")

                c.setFont("Helvetica-Bold", 8.4)
                c.setFillColorRGB(0.02, 0.5, 0.33)
                c.drawRightString(x + label_w - pad_x, price_y + 12, badge)
            elif price:
                c.setFont("Helvetica-Bold", 20.5)
                c.setFillColorRGB(0.08, 0.1, 0.13)
                c.drawCentredString(x + label_w / 2, price_y, price)
            else:
                c.setFont("Helvetica", 8)
                c.setFillColorRGB(0.35, 0.4, 0.45)
                c.drawCentredString(x + label_w / 2, price_y + 3, "Prix à renseigner")

        # Secondary info line (compact)
        extras = []
        if "price_unit" in fields:
            per_unit, suffix = _price_per_unit(product, currency_code=currency_code)
            if per_unit is not None and suffix:
                extras.append(f"{per_unit:.2f} {suffix}")
        pack_info = _format_pack_info(product)
        if pack_info:
            extras.append(pack_info)

        if extras:
            c.setFont("Helvetica", 7.2)
            c.setFillColorRGB(0.35, 0.4, 0.45)
            c.drawRightString(x + label_w - pad_x, price_y - 10, _truncate_text(c, " · ".join(extras), "Helvetica", 7.2, label_w - 2 * pad_x))

        # Two small metadata lines (left)
        meta_lines = []
        if "tva" in fields and product.tva is not None:
            meta_lines.append(f"TVA {product.tva}%")
        if "supplier" in fields and product.supplier:
            meta_lines.append(f"Fourn. {product.supplier}")
        if "brand" in fields and product.brand:
            meta_lines.append(f"Marque {product.brand}")
        if "dlc" in fields:
            label_txt, date_txt = _format_label_date(product)
            if label_txt and date_txt:
                meta_lines.append(f"{label_txt} {date_txt}")

        meta_lines = meta_lines[:2]
        if meta_lines:
            c.setFont("Helvetica", 7.1)
            c.setFillColorRGB(0.35, 0.4, 0.45)
            yy = price_y - 10
            # put on left, avoid overlap with right extras
            for line in meta_lines:
                c.drawString(x + pad_x, yy, _truncate_text(c, line, "Helvetica", 7.1, label_w * 0.62))
                yy -= 8.0

        # Barcode
        barcode_obj.drawOn(c, barcode_x, barcode_y)

        # Code under barcode
        c.setFont("Helvetica", 6.8)
        c.setFillColorRGB(0.35, 0.4, 0.45)
        c.drawCentredString(x + label_w / 2, barcode_y - 7, _truncate_text(c, code, "Helvetica", 6.8, label_w - 2 * pad_x))

        return True

    index = 0
    for product in label_entries:
        col = index % cols
        row = (index // cols) % rows
        x = margin_x + col * label_w
        y = height - margin_y - (row + 1) * label_h

        if not draw_label(x, y, product):
            index += 1
            continue

        index += 1
        if index % (cols * rows) == 0:
            c.showPage()

    c.save()
    pdf_bytes = buffer.getvalue()

    _log_labels_pdf_event(
        tenant=tenant,
        user=request.user,
        params={"count": len(label_entries), "service": service_param},
    )
    logger.info(
        "labels_pdf_generated",
        extra={
            "tenant_id": tenant.id,
            "service": service_param,
            "count": len(label_entries),
        },
    )

    resp = Response(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="stockscan_labels.pdf"'
    return resp
